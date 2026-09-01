from __future__ import annotations

import asyncio
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.enterprise_capabilities.artifacts.service import table_generate
from app.enterprise_capabilities.runtime import CapabilityExecutionContext, InternalCapabilityCatalog
from app.enterprise_capabilities.spreadsheets import WorkbookContractError, normalize_workbook_spec
from app.services.spreadsheets import SpreadsheetService


MODEL_TABLE_PAYLOAD = {
    "sheets": [{
        "name": "经营汇总",
        "tables": [{
            "name": "汇总",
            "headers": ["门店", "销售目标_万元", "实际销售额_万元", "订单数", "销售完成率", "销售缺口_万元"],
            "rows": [
                ["A店", 100, 90, 10, 0.9, 10],
                ["B店", 200, 150, 15, 0.75, 50],
                ["C店", 100, 60, 5, 0.6, 40],
                ["合计", 400, 300, 30, 0.75, 100],
            ],
            "freeze": "A2",
            "formats": {"number": {
                "销售目标_万元": "0.00",
                "实际销售额_万元": "0.00",
                "订单数": "0",
                "销售完成率": "0.00%",
                "销售缺口_万元": "0.00",
            }},
        }],
    }],
}


def _context() -> CapabilityExecutionContext:
    return CapabilityExecutionContext(
        tenant_id="tenant-a",
        user_id="user-a",
        conversation_id="conversation-a",
        kernel_session_id="session-a",
        profile_version="profile-a",
        action_id="action-a",
    )


def test_common_model_table_shape_is_normalized_without_data_loss() -> None:
    spec = normalize_workbook_spec(MODEL_TABLE_PAYLOAD)
    sheet = spec.sheets[0]
    assert [column.title for column in sheet.columns] == MODEL_TABLE_PAYLOAD["sheets"][0]["tables"][0]["headers"]
    assert [row[sheet.columns[0].key] for row in sheet.rows] == ["A店", "B店", "C店", "合计"]
    assert sheet.columns[4].type == "percent"
    assert sheet.columns[4].number_format == "0.00%"
    assert sheet.formatting["freeze_header"] is True


def test_empty_or_unknown_workbook_is_rejected_instead_of_rendered() -> None:
    with pytest.raises(WorkbookContractError, match="column"):
        normalize_workbook_spec({"sheets": [{"name": "empty", "tables": [{"rows": [[1]]}]}]})
    with pytest.raises(WorkbookContractError, match="data row"):
        normalize_workbook_spec({
            "workbook_title": "empty",
            "sheets": [{"name": "empty", "columns": [{"key": "a", "title": "A"}], "rows": []}],
        })


def test_renderer_produces_real_cells_and_number_formats() -> None:
    spec = normalize_workbook_spec(MODEL_TABLE_PAYLOAD)
    content = SpreadsheetService().render_bytes(spec)
    workbook = load_workbook(BytesIO(content), data_only=False)
    sheet = workbook["经营汇总"]
    assert sheet.max_row == 5
    assert sheet.max_column == 6
    assert sheet["A2"].value == "A店"
    assert sheet["B5"].value == 400
    assert sheet["E2"].value == 0.9
    assert sheet["E2"].number_format == "0.00%"
    assert sheet.freeze_panes == "A2"
    workbook.close()


def test_dsh_schema_advertises_canonical_populated_workbook_contract() -> None:
    definition = next(
        item for item in InternalCapabilityCatalog().definitions()
        if item.capability_ref == "artifact.table_generate@v1"
    )
    workbook = definition.input_schema["properties"]["workbook"]
    sheet = workbook["properties"]["sheets"]["items"]
    assert workbook["required"] == ["workbook_title", "sheets"]
    assert set(sheet["required"]) == {"name", "columns", "rows"}
    assert sheet["properties"]["columns"]["minItems"] == 1
    assert sheet["properties"]["rows"]["minItems"] == 1
    assert "tables" not in sheet["properties"]
    assert set(definition.input_schema["required"]) == {"workbook", "delivery_scope"}


def test_table_generate_returns_verified_workbook_summary(monkeypatch) -> None:
    captured = {}

    async def fake_render(spec, *, user_id, filename=None):
        captured["spec"] = spec
        return {
            "type": "xlsx",
            "filename": filename,
            "object_path": f"{user_id}/{filename}",
            "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "sheet_count": 1,
            "validation": {
                "verified": True,
                "sheet_count": 1,
                "sheets": [{"name": "经营汇总", "column_count": 6, "data_row_count": 4}],
            },
        }

    monkeypatch.setattr(
        "app.enterprise_capabilities.artifacts.service.spreadsheet_service.render",
        fake_render,
    )
    result = asyncio.run(table_generate(
        {"filename": "门店经营汇总.xlsx", "workbook": MODEL_TABLE_PAYLOAD},
        _context(),
    ))
    assert result["success"] is True
    assert result["artifact"]["validation"]["verified"] is True
    assert result["artifact"]["validation"]["sheets"][0]["data_row_count"] == 4
    assert len(captured["spec"].sheets[0].rows) == 4


def test_table_generate_hides_declared_intermediate_workbook(monkeypatch) -> None:
    async def fake_render(_spec, *, user_id, filename=None):
        return {"type": "xlsx", "filename": filename, "object_path": f"{user_id}/{filename}"}

    monkeypatch.setattr("app.enterprise_capabilities.artifacts.service.spreadsheet_service.render", fake_render)
    result = asyncio.run(table_generate({
        "filename": "handoff.xlsx", "workbook": MODEL_TABLE_PAYLOAD,
        "delivery_scope": "intermediate",
    }, _context()))
    assert result["artifact"]["lifecycle"] == "intermediate"
    assert result["artifact"]["visibility"] == "internal"
