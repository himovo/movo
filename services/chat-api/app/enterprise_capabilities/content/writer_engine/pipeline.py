from __future__ import annotations
from app.infrastructure.observability.config import log_print

import json
import re
from types import SimpleNamespace
from typing import Any, AsyncIterator, Dict, List

from app.llm.factory import get_request_scoped_llm_client
from app.llm.types import Message, Role
from app.enterprise_capabilities.content.body_projection import body_only_contract_context, strip_visual_placeholder_directives
from app.enterprise_capabilities.content.evaluation import (
    Standard,
    issues_as_writer_feedback,
    issues_summary,
)
from app.enterprise_capabilities.content.evaluation.issue_finder import IssueFinder
from app.enterprise_capabilities.evidence.foundation.writer_packet import WriterEvidencePacket, build_writer_packet_source_references
from app.enterprise_capabilities.content.evaluation.standards_generator import StandardsGenerator
from app.tools.infographic import generate_infographic_asset, is_valid_remote_image_url
from app.enterprise_capabilities.content.writer_engine.visual_budget import resolve_sectional_visual_budget
from app.enterprise_capabilities.content.writer_engine.commentary_relay import CommentaryRelay
from app.enterprise_capabilities.content.writer_engine.unified_compose.components import (
    ComposeDeps,
    MarkdownPostProcessor,
    ObservationAdapter,
    OutlinePlanner,
    SectionWriter,
    SinglePassWriter,
    VisualAugmenter,
)
from app.enterprise_capabilities.content.writer_engine.unified_compose.prompt_registry import PromptRegistry


class WriterEnginePipeline:
    """Bottom-layer writing pipeline with fixed structured flow.

    This pipeline always runs: outline -> section writing -> visual planning -> post-process.
    """

    def __init__(self) -> None:
        self._llm = get_request_scoped_llm_client(streaming=False, stage="compose", intent="generation")
        self._prompts = PromptRegistry()
        deps = ComposeDeps(llm=self._llm, language_name=self._language_name, log=self._log_stage)
        self._planner = OutlinePlanner(deps, self._prompts)
        self._writer = SectionWriter(deps, self._prompts)
        self._visual = VisualAugmenter(deps, self._prompts)
        self._single = SinglePassWriter(deps, self._prompts)
        # Lazily-initialized — only constructed when the dynamic-standards
        # flag is on and a sectional run actually triggers evaluation.
        self._standards_generator: StandardsGenerator | None = None
        self._issue_finder: IssueFinder | None = None

    def _eval_components(self) -> tuple[StandardsGenerator, IssueFinder]:
        if self._standards_generator is None:
            self._standards_generator = StandardsGenerator()
        if self._issue_finder is None:
            self._issue_finder = IssueFinder()
        return self._standards_generator, self._issue_finder

    @staticmethod
    def _language_name(language: str) -> str:
        return "中文" if language == "zh" else "English"

    @staticmethod
    def _amend_with_doc_feedback(user_query: str, *, strategy: Dict[str, Any] | None = None, output_spec: Dict[str, Any] | None = None) -> str:
        """Append document-level review feedback to user_query if present.

        This is the retry-feedback channel for the document-level dynamic
        evaluation: when task_satisfaction_eval finds verdict != pass, it
        calls _retry_after_finalize and writes formatted feedback into
        output_spec (and tool_writer_engine_compose mirrors it onto
        strategy). All generate methods call this helper at their entry,
        so the LLM sees the previous run's issues as additional input.
        """
        feedback = ""
        if isinstance(strategy, dict):
            feedback = str(strategy.get("doc_level_feedback") or "").strip()
        if not feedback and isinstance(output_spec, dict):
            feedback = str(output_spec.get("__doc_level_eval_feedback") or "").strip()
        if not feedback:
            return user_query
        try:
            log_print(
                "[writer_engine][doc_feedback] applied len=%d preview=%r"
                % (len(feedback), feedback[:120]),
                flush=True,
            )
        except Exception:
            pass
        return f"{user_query}\n\n{feedback}"

    @staticmethod
    def _log_stage(code: str, meta: Dict[str, Any] | None = None) -> None:
        try:
            meta = dict(meta or {})
            if code == "outline_prompt":
                log_print(
                    "[writer_engine] outline prompt | target_sections=%s depth=%s"
                    % (
                        int(meta.get("target_sections") or 0),
                        int(meta.get("outline_depth") or 0),
                    ),
                    flush=True,
                )
                return
            if code == "outline_result":
                log_print(
                    "[writer_engine] outline result | title=%s raw=%s used=%s target=%s"
                    % (
                        str(meta.get("title") or "")[:120],
                        int(meta.get("section_count_raw") or 0),
                        int(meta.get("section_count_used") or 0),
                        int(meta.get("target_sections") or 0),
                    ),
                    flush=True,
                )
                return
            if code == "outline_alignment_report":
                log_print(
                    "[writer_engine] outline alignment | input=%s output=%s target=%s required=%s"
                    % (
                        int(meta.get("input_count") or 0),
                        int(meta.get("output_count") or 0),
                        int(meta.get("target_count") or 0),
                        len(list(meta.get("required_blocks") or [])),
                    ),
                    flush=True,
                )
                return
            if code == "section_prompt":
                log_print(
                    "[writer_engine] section prompt | number=%s title=%s target_words=%s"
                    % (
                        str(meta.get("section_number") or ""),
                        str(meta.get("section_title") or "")[:120],
                        int(meta.get("target_words") or 0),
                    ),
                    flush=True,
                )
                return
            if code == "visual_prompt":
                log_print(
                    "[writer_engine] visual prompt | section=%s"
                    % (str(meta.get("section_title") or "")[:120],),
                    flush=True,
                )
                return
            log_print(
                "[writer_engine][event] code=%s meta=%s"
                % (code, json.dumps(meta, ensure_ascii=False)),
                flush=True,
            )
        except Exception:
            pass

    @staticmethod
    def _section_word_target(strategy: Dict[str, Any], section_count: int, idx: int) -> int:
        wb = strategy.get("word_budget") if isinstance(strategy.get("word_budget"), dict) else {}
        total_target = int(wb.get("total_target") or 2600)
        section_min = int(wb.get("section_min") or 240)
        section_max = int(wb.get("section_max") or 520)
        if section_count <= 0:
            section_count = 1
        base = max(section_min, min(section_max, total_target // section_count))
        if idx in {0, section_count - 1}:
            base = max(section_min, int(base * 0.9))
        return max(section_min, min(section_max, base))

    @staticmethod
    def _build_references_markdown(refs: List[Dict[str, Any]], language: str) -> str:
        if not refs:
            return ""
        head = "## 参考来源\n" if language == "zh" else "## References\n"
        lines = [head]
        for item in refs[:30]:
            idx = int(item.get("index") or 0)
            url = str(item.get("url") or "").strip()
            if not url:
                continue
            title = str(item.get("title") or "").strip() or url
            lines.append(f"{idx}. [{title}]({url})")
        return "\n".join(lines).strip()

    @staticmethod
    def _compact_running_summary(text: str, limit: int = 2200) -> str:
        raw = (text or "").strip()
        if len(raw) <= limit:
            return raw
        return raw[-limit:]

    async def _build_section_infographic_prompt(
        self,
        *,
        section_title: str,
        section_markdown: str,
        language: str,
    ) -> Dict[str, str]:
        payload = {
            "section_title": section_title,
            "section_excerpt": section_markdown[:2200],
            "language": language,
            "constraints": [
                "single infographic image",
                "clean business layout",
                "readable labels",
                "no watermark",
                "all visible text must be reader-facing final copy",
                "no process labels like 强结论开头/导语/正文/结尾/section/introduction/conclusion",
                "no editorial or writing-instruction text inside the image",
            ],
        }
        system = self._prompts.get("visual_prompt_builder")
        try:
            from app.llm.types import Message, Role

            resp = await self._llm.ainvoke(
                [Message(role=Role.SYSTEM, content=system), Message(role=Role.USER, content=json.dumps(payload, ensure_ascii=False))]
            )
            raw = str(getattr(resp, "content", "") or "").strip()
            data = json.loads(raw.replace("```json", "").replace("```", "").strip())
            if isinstance(data, dict):
                prompt = str(data.get("prompt") or "").strip()
                if prompt:
                    return {
                        "prompt": prompt,
                        "negative_prompt": str(data.get("negative_prompt") or "").strip(),
                        "size": str(data.get("size") or "1664*928").strip(),
                    }
        except Exception as e:
            self._log_stage("infographic_prompt_failed", {"error": str(e)[:260]})
        fallback_prompt = (
            ("中文信息图，主题：%s。仅使用面向最终读者的标题、标签和图注文案。不要出现“强结论开头”“导语”“正文”“结尾”等过程性或结构性文字。突出关键结论、驱动因素、风险与建议，布局为标题+三栏要点+结论区，风格专业简洁，文字清晰可读。")
            % (section_title or "章节分析")
            if language == "zh"
            else (
                f"Professional infographic about '{section_title}'. Use only reader-facing final labels and captions. "
                "Do not include process or structural words like opening, introduction, section, conclusion, draft, or outline. "
                "Emphasize key findings, drivers, risks, and recommendations. Layout: title + three insight columns + conclusion strip. Clear readable labels."
            )
        )
        return {
            "prompt": fallback_prompt,
            "negative_prompt": "low resolution, blurry text, distorted layout, watermark, messy composition, editorial labels, process text, outline text, section heading text, draft wording",
            "size": "1664*928",
        }

    async def _maybe_generate_section_infographic(
        self,
        *,
        section_no: str,
        section_title: str,
        section_markdown: str,
        language: str,
        user_id: str,
    ) -> Dict[str, Any]:
        if self._count_real_image_urls(section_markdown) > 0:
            return {"ok": False, "reason": "section_has_image_already"}
        prompt_data = await self._build_section_infographic_prompt(
            section_title=section_title,
            section_markdown=section_markdown,
            language=language,
        )
        prompt = str(prompt_data.get("prompt") or "").strip()
        negative_prompt = str(prompt_data.get("negative_prompt") or "").strip()
        size = str(prompt_data.get("size") or "1664*928").strip()
        self._log_stage("infographic_prompt", {"section": section_no, "title": section_title[:120], "size": size})
        result = await generate_infographic_asset(
            prompt=prompt,
            negative_prompt=negative_prompt,
            size=size,
            user_id=str(user_id or "anonymous"),
        )
        image_url = str(result.get("image_url") or "").strip()
        ok = bool(result.get("ok")) and is_valid_remote_image_url(image_url)
        self._log_stage("infographic_result", {"section": section_no, "ok": ok, "url": image_url[:220]})
        if not ok:
            return {"ok": False, "reason": str(result.get("error") or "generate_failed")}
        return {
            "ok": True,
            "url": image_url,
            "caption": section_title,  # Use section title directly, not editorial marks
        }

    @staticmethod
    def _count_markdown_images(content: str) -> int:
        text = str(content or "")
        if not text:
            return 0
        return len(re.findall(r"!\[[^\]]*\]\(([^)]+)\)", text))

    @staticmethod
    def _count_real_image_urls(content: str) -> int:
        text = str(content or "")
        if not text:
            return 0
        urls = re.findall(r"!\[[^\]]*\]\(([^)]+)\)", text)
        real = [u for u in urls if is_valid_remote_image_url(str(u).strip())]
        return len(real)

    @staticmethod
    def _fill_non_http_image_links(content: str, urls: List[str]) -> tuple[str, int]:
        text = str(content or "")
        candidates = [str(u).strip() for u in (urls or []) if is_valid_remote_image_url(str(u).strip())]
        if not text or not candidates:
            return text, 0
        used = 0

        def _repl(m: re.Match[str]) -> str:
            nonlocal used
            alt = str(m.group(1) or "").strip() or f"image_{used+1}"
            link = str(m.group(2) or "").strip()
            if is_valid_remote_image_url(link):
                return m.group(0)
            if used >= len(candidates):
                return m.group(0)
            new_link = candidates[used]
            used += 1
            return f"![{alt}]({new_link})"

        rewritten = re.sub(r"!\[([^\]]*)\]\(([^)]*)\)", _repl, text)
        return rewritten, used

    @staticmethod
    def _extract_image_slots(content: str, *, user_query: str, language: str) -> tuple[str, List[Dict[str, str]]]:
        text = str(content or "")
        if not text.strip():
            return text, []
        slots: List[Dict[str, str]] = []
        idx = 0

        def _repl(m: re.Match[str]) -> str:
            nonlocal idx
            idx += 1
            body = str(m.group(1) or "")
            alt = ""
            prompt = ""
            for line in body.splitlines():
                s = str(line or "").strip()
                if not s:
                    continue
                if s.lower().startswith("alt:"):
                    alt = s.split(":", 1)[1].strip()
                elif s.lower().startswith("prompt:"):
                    prompt = s.split(":", 1)[1].strip()
            if not alt:
                alt = f"{'信息图' if language == 'zh' else 'Infographic'} {idx}"
            if not prompt:
                prompt = (
                    f"生成一张与文章内容相关的信息图，主题：{user_query}。突出关键概念、流程、应用场景，文字清晰可读。"
                    if language == "zh"
                    else f"Create one infographic for: {user_query}. Emphasize core concepts, workflow, and use cases."
                )
            token = f"[[IMAGE_SLOT_{idx}]]"
            slots.append({"token": token, "alt": alt, "prompt": prompt})
            return token

        rewritten = re.sub(r"\[IMAGE_SLOT\]\s*([\s\S]*?)\s*\[/IMAGE_SLOT\]", _repl, text, flags=re.IGNORECASE)
        return rewritten, slots

    @staticmethod
    def _build_planned_image_slots(
        *,
        content: str,
        user_query: str,
        strategy: Dict[str, Any],
        language: str,
    ) -> tuple[str, List[Dict[str, str]]]:
        text = str(content or "")
        if not text.strip():
            return text, []
        content_plan = strategy.get("content_plan") if isinstance(strategy.get("content_plan"), dict) else {}
        plan_sections = [item for item in list(content_plan.get("sections") or []) if isinstance(item, dict)]
        visual_slots = [item for item in list(content_plan.get("visual_slots") or []) if isinstance(item, dict)]
        if not visual_slots:
            return text, []

        section_title_by_id = {
            str(item.get("section_id") or "").strip(): str(item.get("title") or "").strip()
            for item in plan_sections
            if str(item.get("section_id") or "").strip()
        }
        section_order_by_id = {
            str(item.get("section_id") or "").strip(): idx
            for idx, item in enumerate(plan_sections)
            if str(item.get("section_id") or "").strip()
        }

        def _norm_title(value: str) -> str:
            raw = str(value or "").strip().lower()
            raw = re.sub(r"^\d+(?:\.\d+)*\s*", "", raw)
            raw = re.sub(r"[^\w\u4e00-\u9fff]+", "", raw)
            return raw

        def _heading_matches(heading_line: str, anchor_title: str) -> bool:
            h = _norm_title(heading_line)
            a = _norm_title(anchor_title)
            if not h or not a:
                return False
            return a in h or h in a

        slot_defs: List[Dict[str, str]] = []

        for idx, slot in enumerate(visual_slots, start=1):
            role = str(slot.get("role") or "visual").strip()
            anchor_id = str(slot.get("anchor_section_id") or "").strip()
            section_title = section_title_by_id.get(anchor_id, "")
            desc = str(slot.get("description") or "").strip()
            alt = section_title or desc or (f"{'信息图' if language == 'zh' else 'Infographic'} {idx}")
            prompt = (
                f"Create one concise infographic for the article topic: {user_query}. "
                f"Visual role: {role}. "
                f"Anchor section: {section_title or 'general context'}. "
                f"Focus: {desc or 'highlight the section’s core idea with readable labels and clean structure'}."
            )
            slot_defs.append(
                {
                    "token": f"[[PLANNED_IMAGE_SLOT_{idx}]]",
                    "alt": alt[:120],
                    "prompt": prompt[:900],
                    "anchor_title": section_title,
                    "anchor_order": int(section_order_by_id.get(anchor_id, len(slot_defs))),
                }
            )

        rewritten = text
        for slot in slot_defs:
            token = str(slot.get("token") or "").strip()
            anchor_title = str(slot.get("anchor_title") or "").strip()
            inserted = False
            current_headings = list(re.finditer(r"(?m)^(#{2,6}\s+.+)$", rewritten))
            if anchor_title and current_headings:
                for match in current_headings:
                    heading_line = str(match.group(1) or "")
                    if _heading_matches(heading_line, anchor_title):
                        insert_at = match.end()
                        rewritten = rewritten[:insert_at] + f"\n\n{token}\n" + rewritten[insert_at:]
                        inserted = True
                        break
            if not inserted and current_headings:
                order = max(0, min(int(slot.get("anchor_order") or 0), len(current_headings) - 1))
                match = current_headings[order]
                insert_at = match.end()
                rewritten = rewritten[:insert_at] + f"\n\n{token}\n" + rewritten[insert_at:]
                inserted = True
            if not inserted:
                rewritten = rewritten.rstrip() + f"\n\n{token}\n"
        return rewritten, [{"token": s["token"], "alt": s["alt"], "prompt": s["prompt"]} for s in slot_defs]

    async def _materialize_image_slots(
        self,
        *,
        content: str,
        slots: List[Dict[str, str]],
        user_id: str,
        language: str,
    ) -> tuple[str, List[str]]:
        text = str(content or "")
        used_urls: List[str] = []
        for i, slot in enumerate(slots, start=1):
            token = str(slot.get("token") or "").strip()
            if not token:
                continue
            prompt = str(slot.get("prompt") or "").strip()
            alt = str(slot.get("alt") or "").strip() or (f"{'信息图' if language == 'zh' else 'Infographic'} {i}")
            if not prompt:
                text = text.replace(token, "")
                continue
            try:
                result = await generate_infographic_asset(
                    prompt=prompt,
                    user_id=str(user_id or "anonymous"),
                    size="1664*928",
                )
                url = str(result.get("image_url") or "").strip()
                ok = is_valid_remote_image_url(url)
                self._log_stage("image_slot_result", {"slot": i, "ok": ok, "url": url[:220]})
                if ok:
                    used_urls.append(url)
                    text = text.replace(token, f"![{alt}]({url})")
                else:
                    text = text.replace(token, "")
            except Exception as e:
                self._log_stage("image_slot_failed", {"slot": i, "error": str(e)[:220]})
                text = text.replace(token, "")
        return text, [u for j, u in enumerate(used_urls) if u and u not in used_urls[:j]]

    async def _generate_fallback_infographic_urls(self, user_goal: str, count: int, user_id: str) -> List[str]:
        n = max(0, min(int(count or 0), 6))
        if n <= 0:
            return []
        base = str(user_goal or "").strip() or "technical article"
        urls: List[str] = []
        for idx in range(n):
            prompt = (
                "Generate a clean technology infographic for the current document. "
                f"Theme: {base}. Variant: {idx + 1}. "
                "Use concise labels and readable layout."
            )
            try:
                result = await generate_infographic_asset(
                    prompt=prompt,
                    user_id=str(user_id or "anonymous"),
                    size="1664*928",
                )
                u = str(result.get("image_url") or "").strip()
                if is_valid_remote_image_url(u):
                    urls.append(u)
            except Exception:
                continue
        return [u for i, u in enumerate(urls) if u and u not in urls[:i]]

    async def generate_direct(
        self,
        *,
        user_query: str,
        language: str,
        strategy: Dict[str, Any],
        tool_observations: List[Dict[str, Any]],
        selected_style_md: str,
        evidence_bundle: Dict[str, Any],
        writer_evidence_packet: Dict[str, Any] | None = None,
        user_id: str = "anonymous",
    ) -> AsyncIterator[Dict[str, Any]]:
        user_query = self._amend_with_doc_feedback(user_query, strategy=strategy)
        pseudo_context = SimpleNamespace(
            intent="generation",
            payload={
                "compose_strategy": dict(strategy or {}),
                "tool_observations": list(tool_observations or []),
                "evidence_bundle": dict(evidence_bundle or {}),
                "writer_evidence_packet": dict(writer_evidence_packet or {}),
            },
            output_spec={},
        )
        yield {
            "type": "activity",
            "content": {
                "kind": "writing",
                "message": (
                    "正在一次性撰写完整正文"
                    if language == "zh" else "Writing the complete draft in one pass"
                ),
            },
        }
        commentary_relay: CommentaryRelay[str] = CommentaryRelay()
        operation = self._single.write_single_pass(
            context=pseudo_context,
            language=language,
            user_query=user_query,
            tool_observations=tool_observations,
            selected_style_md=selected_style_md,
            commentary_sink=commentary_relay.publish,
        )
        async for payload in commentary_relay.stream(operation):
            yield {"type": "commentary", "content": payload}
        content = commentary_relay.result
        content = MarkdownPostProcessor.sanitize_markdown_noise(content)
        content = MarkdownPostProcessor.normalize_list_item_style(content)
        content = MarkdownPostProcessor.sanitize_runtime_artifacts(str(content or ""))
        yield {"type": "answer", "content": (content or "").strip() + "\n\n"}

        refs = self._build_references_markdown(ObservationAdapter.build_source_index(tool_observations), language)
        if refs:
            yield {"type": "answer", "content": refs + "\n"}

        self._log_stage(
            "direct_report_metrics",
            {
                "mode": str((strategy or {}).get("mode") or "direct_compose"),
                "total_len": len(content or ""),
                "has_refs": bool(refs),
            },
        )

    @staticmethod
    def _normalize_sections_from_contract(structure_contract: Dict[str, Any]) -> List[Dict[str, Any]]:
        if not isinstance(structure_contract, dict):
            return []
        raw_blocks = structure_contract.get("blocks")
        if not isinstance(raw_blocks, list):
            return []
        out: List[Dict[str, Any]] = []
        for item in raw_blocks:
            if not isinstance(item, dict):
                continue
            title = str(item.get("title") or item.get("name") or item.get("block_name") or "").strip()
            if not title:
                continue
            out.append(
                {
                    "title": title,
                    "role": str(item.get("role") or "").strip(),
                    "objective": str(item.get("objective") or item.get("purpose") or "").strip(),
                    "key_points": [str(x).strip() for x in (item.get("key_points") or []) if str(x).strip()][:8],
                    "evidence_focus": [str(x).strip() for x in (item.get("evidence_focus") or []) if str(x).strip()][:8],
                    "visual_hint": str(item.get("visual_hint") or "").strip().lower(),
                }
            )
        return out

    @staticmethod
    def _merge_sections_with_body_plan(
        *,
        sections: List[Dict[str, Any]],
        body_plan: Dict[str, Any],
    ) -> List[Dict[str, Any]]:
        body_sections = [item for item in list((body_plan or {}).get("sections") or []) if isinstance(item, dict)]
        if not sections or not body_sections:
            return list(sections or [])

        def _norm_title(value: Any) -> str:
            title = str(value or "").strip().lower()
            title = re.sub(r"^\d+(?:\.\d+)*\s+", "", title)
            title = re.sub(r"\s+", " ", title)
            return title

        body_by_title: Dict[str, List[Dict[str, Any]]] = {}
        for item in body_sections:
            key = _norm_title(item.get("title"))
            if not key:
                continue
            bucket = body_by_title.setdefault(key, [])
            bucket.append(item)

        consumed_ids = set()
        merged_sections: List[Dict[str, Any]] = []
        for idx, section in enumerate(list(sections or [])):
            merged = dict(section or {})
            title_key = _norm_title(merged.get("title"))
            source: Dict[str, Any] | None = None
            candidates = body_by_title.get(title_key) or []
            for candidate in candidates:
                candidate_id = str(candidate.get("section_id") or "")
                if candidate_id and candidate_id in consumed_ids:
                    continue
                source = candidate
                break
            if source is None and idx < len(body_sections):
                candidate = body_sections[idx]
                candidate_id = str(candidate.get("section_id") or "")
                if not candidate_id or candidate_id not in consumed_ids:
                    source = candidate
            if source is None:
                merged_sections.append(merged)
                continue

            source_id = str(source.get("section_id") or "")
            if source_id:
                consumed_ids.add(source_id)
                merged["section_id"] = source_id

            source_objective = str(source.get("purpose") or source.get("objective") or "").strip()
            current_objective = str(merged.get("objective") or "").strip()
            source_role = str(source.get("role") or "").strip()
            if source_role:
                merged["role"] = source_role
            if source_objective and (
                not current_objective
                or current_objective in {
                    "围绕现有事实组织内容",
                    "Organize the available facts for this section",
                    "补充该核心内容的论述",
                    "Cover specific anchor section",
                }
            ):
                merged["objective"] = source_objective

            source_key_points = [str(x).strip() for x in list(source.get("key_points") or []) if str(x).strip()][:8]
            if source_key_points and not list(merged.get("key_points") or []):
                merged["key_points"] = source_key_points

            source_evidence_focus = [
                str(x).strip() for x in list(source.get("evidence_focus") or []) if str(x).strip()
            ][:8]
            if source_evidence_focus and not list(merged.get("evidence_focus") or []):
                merged["evidence_focus"] = source_evidence_focus

            for key, limit in (
                ("allowed_claim_types", 8),
                ("disallowed_claim_types", 8),
                ("open_questions", 8),
            ):
                source_values = [str(x).strip() for x in list(source.get(key) or []) if str(x).strip()][:limit]
                if source_values and not list(merged.get(key) or []):
                    merged[key] = source_values
            source_must_cover_facts = WriterEnginePipeline._normalize_section_facts(source.get("must_cover_facts"))[:12]
            if source_must_cover_facts and not list(merged.get("must_cover_facts") or []):
                merged["must_cover_facts"] = source_must_cover_facts

            source_visual_hint = str(source.get("visual_hint") or "").strip().lower()
            if source_visual_hint and str(merged.get("visual_hint") or "").strip().lower() in {"", "none"}:
                merged["visual_hint"] = source_visual_hint

            source_argument_items = [item for item in list(source.get("argument_items") or []) if isinstance(item, dict)]
            if source_argument_items and not list(merged.get("argument_items") or []):
                merged["argument_items"] = source_argument_items[:8]

            source_target_words = int(source.get("target_words") or 0)
            if source_target_words > 0 and int(merged.get("target_words") or 0) <= 0:
                merged["target_words"] = source_target_words

            merged_sections.append(merged)
        return merged_sections

    def _align_outline_sections(
        self,
        *,
        sections: List[Dict[str, Any]],
        section_count: int,
        structure_contract: Dict[str, Any],
        language: str,
    ) -> List[Dict[str, Any]]:
        normalized = list(sections or [])
        required_blocks = [str(x).strip() for x in (structure_contract.get("required_blocks") or []) if str(x).strip()]
        target_count = len(required_blocks) if required_blocks else int(section_count or 1)
        target_count = max(1, min(24, target_count))

        def _hint(v: str) -> str:
            raw = str(v or "").strip().lower()
            mapping = {
                "markdown_table": "table",
                "md_table": "table",
                "ascii_diagram": "mermaid",
                "callout_blockquote": "none",
                "blockquote": "none",
                "list": "none",
            }
            raw = mapping.get(raw, raw)
            return raw if raw in {"table", "chart", "mermaid", "none"} else "none"

        used = [False] * len(normalized)
        aligned: List[Dict[str, Any]] = []

        # 1) Strictly cover required anchors first
        for req in required_blocks:
            req_lower = req.lower()
            hit_idx = -1
            for i, sec in enumerate(normalized):
                if used[i]:
                    continue
                title_lower = str(sec.get("title") or "").strip().lower()
                if req_lower in title_lower or title_lower in req_lower:
                    hit_idx = i
                    break
            if hit_idx >= 0:
                used[hit_idx] = True
                matched = dict(normalized[hit_idx])
                matched["title"] = req
                aligned.append(matched)
            else:
                aligned.append(
                    {
                        "title": req,
                        "objective": "补充该核心内容的论述" if language == "zh" else "Cover specific anchor section",
                        "key_points": [],
                        "evidence_focus": [],
                        "visual_hint": "none",
                    }
                )

        # 2) Fill from remaining model sections
        if len(aligned) < target_count:
            for i, sec in enumerate(normalized):
                if len(aligned) >= target_count:
                    break
                if used[i]:
                    continue
                aligned.append(dict(sec))
                used[i] = True

        # 3) Pad placeholders if still short
        while len(aligned) < target_count:
            n = len(aligned) + 1
            aligned.append(
                {
                    "title": f"{'章节' if language == 'zh' else 'Section'} {n}",
                    "objective": "",
                    "key_points": [],
                    "evidence_focus": [],
                    "visual_hint": "none",
                }
            )

        # 4) Strict truncate to target_count
        aligned = aligned[:target_count]

        # Fix numbering and defaults
        for idx, sec in enumerate(aligned):
            sec["number"] = str(idx + 1)
            sec["level"] = 1
            if not sec.get("title"):
                sec["title"] = f"{'章节' if language == 'zh' else 'Section'} {idx + 1}"
            sec["key_points"] = [str(x) for x in (sec.get("key_points") or [])][:8]
            sec["evidence_focus"] = [str(x) for x in (sec.get("evidence_focus") or [])][:8]
            sec["visual_hint"] = _hint(sec.get("visual_hint") or "none")

        return aligned

    async def generate(
        self,
        *,
        user_query: str,
        language: str,
        strategy: Dict[str, Any],
        tool_observations: List[Dict[str, Any]],
        selected_style_md: str,
        evidence_bundle: Dict[str, Any],
        writer_evidence_packet: Dict[str, Any],
        user_id: str = "anonymous",
    ) -> AsyncIterator[Dict[str, Any]]:
        user_query = self._amend_with_doc_feedback(user_query, strategy=strategy)
        if not writer_evidence_packet:
            raise RuntimeError("writer_evidence_packet is required for sectional compose")
        packet = WriterEvidencePacket.model_validate(writer_evidence_packet or {})
        source_refs = build_writer_packet_source_references(packet)
        self._log_stage(
            "sectional_writer_packet",
            {
                "writer_evidence_packet_active": True,
                "confirmed_facts": len(packet.confirmed_facts),
                "business_datasets": len(packet.business_datasets),
                "computed_metrics": len(packet.computed_metrics),
                "source_excerpts": len(packet.source_excerpts),
                "citations": len(packet.citations),
                "coverage_complete": bool(packet.coverage.complete),
            },
        )
        structure_contract = strategy.get("structure_contract") if isinstance(strategy.get("structure_contract"), dict) else {}
        compose_policy = strategy.get("compose_policy") if isinstance(strategy.get("compose_policy"), dict) else {}
        evidence_policy = strategy.get("evidence_policy") if isinstance(strategy.get("evidence_policy"), dict) else {}
        visual_contract = strategy.get("visual_contract") if isinstance(strategy.get("visual_contract"), dict) else {}
        quality_gates = strategy.get("quality_gates") if isinstance(strategy.get("quality_gates"), dict) else {}
        prompt_contract = strategy.get("prompt_contract") if isinstance(strategy.get("prompt_contract"), dict) else {}
        production_contract = strategy.get("production_contract") if isinstance(strategy.get("production_contract"), dict) else {}
        body_plan = strategy.get("body_plan") if isinstance(strategy.get("body_plan"), dict) else {}

        _mode = str(strategy.get("mode") or "").strip().lower()
        is_agreement = (
            _mode == "agreement_clauses"
            or (bool(strategy.get("skip_toc")) and str(strategy.get("numbering_style") or "").strip().lower() == "clause")
            or bool((evidence_bundle or {}).get("agreement_template_markdown"))
        )
        skip_toc = bool(strategy.get("skip_toc")) or is_agreement
        skip_visuals = bool(strategy.get("skip_visuals")) or is_agreement
        prompt_variant = "agreement" if is_agreement else ""

        section_count = max(1, min(24, int(strategy.get("section_count") or 8)))
        vp = strategy.get("visual_policy") if isinstance(strategy.get("visual_policy"), dict) else {}
        content_task_spec = strategy.get("content_task_spec") if isinstance(strategy.get("content_task_spec"), dict) else {}
        visual_plan = content_task_spec.get("visual_plan") if isinstance(content_task_spec.get("visual_plan"), dict) else {}
        visual_budget = resolve_sectional_visual_budget(
            visual_policy=vp,
            visual_plan=visual_plan,
            section_count=section_count,
        )
        max_visuals = 0 if skip_visuals else max(0, min(8, int(visual_budget.get("max_visuals") or 0)))
        min_visuals = 0 if skip_visuals else max(0, int(visual_budget.get("min_visuals") or 0))
        strategy = dict(strategy or {})
        strategy["section_count"] = section_count
        strategy["numbering_style"] = str(strategy.get("numbering_style") or "hierarchical")
        strategy["outline_depth"] = int(strategy.get("outline_depth") or 2)
        contract_context = {
            "compose_policy": compose_policy,
            "structure_contract": structure_contract,
            "evidence_policy": evidence_policy,
            "visual_contract": visual_contract,
            "quality_gates": quality_gates,
            "prompt_contract": prompt_contract,
            "production_contract": production_contract,
            "forbidden_patterns": list(strategy.get("forbidden_patterns") or []),
        }
        body_contract_context = body_only_contract_context(contract_context)
        try:
            self._log_stage(
                "contract_received",
                {
                    "section_count": section_count,
                    "required_blocks": list(structure_contract.get("required_blocks") or []),
                    "compose_policy_keys": sorted(list(compose_policy.keys())),
                    "structure_keys": sorted(list(structure_contract.keys())),
                    "evidence_keys": sorted(list(evidence_policy.keys())),
                    "visual_keys": sorted(list(visual_contract.keys())),
                    "quality_keys": sorted(list(quality_gates.keys())),
                    "prompt_contract_keys": sorted(list(prompt_contract.keys())),
                    "production_contract_keys": sorted(list(production_contract.keys())),
                },
            )
        except Exception:
            pass

        # Generate evaluation standards once for this sectional run.
        # Standards stay constant across all sections (the user's intent
        # doesn't change mid-document); per-section evaluation reuses the
        # same rubric. If standards generation fails (LLM connection
        # error etc.), the loop below skips section eval gracefully.
        section_standards: List[Standard] = []
        if (user_query or "").strip():
            try:
                gen, _ = self._eval_components()
                section_standards = await gen.generate(user_request=user_query)
                if section_standards:
                    log_print(
                        "[writer_engine][per_section_eval] standards ready count=%d"
                        % len(section_standards),
                        flush=True,
                    )
            except Exception as exc:
                log_print(
                    f"[writer_engine][per_section_eval] standards generation failed: {exc}",
                    flush=True,
                )
                section_standards = []

        yield {
            "type": "activity",
            "content": {
                "kind": "writing",
                "message": "正在规划超长文结构" if language == "zh" else "Planning the ultra-long document structure",
            },
        }
        outline_relay: CommentaryRelay[Dict[str, Any]] = CommentaryRelay()
        outline_operation = self._planner.generate(
            user_query=user_query,
            language=language,
            writer_evidence_packet=packet.model_dump(),
            strategy=strategy,
            style_md=selected_style_md,
            contract_context=body_contract_context,
            prompt_variant=prompt_variant,
            commentary_sink=outline_relay.publish,
        )
        async for payload in outline_relay.stream(outline_operation):
            yield {"type": "commentary", "content": payload}
        outline = outline_relay.result
        title_hint = str(strategy.get("title_hint") or "").strip()
        title = str(title_hint or outline.get("title") or ("结构化文档" if language == "zh" else "Structured Document")).strip()
        sections = list(outline.get("sections") or [])
        if not sections:
            sections = [
                {
                    "number": "1",
                    "level": 1,
                    "title": "概览" if language == "zh" else "Overview",
                    "objective": "",
                    "key_points": [],
                    "evidence_focus": [],
                    "visual_hint": "table",
                }
            ]
        raw_count = len(sections)
        sections = self._align_outline_sections(
            sections=sections,
            section_count=section_count,
            structure_contract=structure_contract,
            language=language,
        )
        sections = self._merge_sections_with_body_plan(sections=sections, body_plan=body_plan)
        try:
            self._log_stage(
                "outline_alignment_report",
                {
                    "input_count": raw_count,
                    "output_count": len(sections),
                    "target_count": max(
                        len([str(x).strip() for x in (structure_contract.get("required_blocks") or []) if str(x).strip()]),
                        int(section_count or 1),
                    ),
                    "required_blocks": [str(x).strip() for x in (structure_contract.get("required_blocks") or []) if str(x).strip()],
                },
            )
        except Exception:
            pass
        if min_visuals > 0:
            self._log_stage("visual_minimum", {"min_visuals": min_visuals, "max_visuals": max_visuals, "sections": len(sections)})

        numbering_style = str(strategy.get("numbering_style") or "hierarchical").strip().lower()
        if skip_toc:
            # A custom no-numbering contract may use its first section as the
            # visual entry point, so avoid injecting a second document title.
            if numbering_style != "none":
                yield {"type": "answer", "content": f"# {title}\n\n"}
        else:
            toc_lines = [f"# {title}", "", "## 目录" if language == "zh" else "## Table of Contents", ""]
            for idx, sec in enumerate(sections, start=1):
                sec_no = str(sec.get("number") or idx)
                sec_title = self._writer.format_display_title(
                    str(sec.get("title") or f"{'章节' if language == 'zh' else 'Section'} {idx}"),
                    sec_no,
                    language,
                    numbering_style,
                )
                toc_entry = sec_title if self._writer._title_has_explicit_numbering(sec_title) else f"{sec_no} {sec_title}".strip()
                toc_lines.append(f"- {toc_entry}")
            yield {"type": "answer", "content": "\n".join(toc_lines).strip() + "\n\n"}

        running_summary = ""
        total = len(sections)
        total_len = 0
        visuals_used = 0
        generated_urls: List[str] = []
        for idx, sec in enumerate(sections):
            sec_no = str(sec.get("number") or (idx + 1))
            sec_level = int(sec.get("level") or 1)
            sec_title = self._writer.format_display_title(
                str(sec.get("title") or f"{'章节' if language == 'zh' else 'Section'} {idx+1}"),
                sec_no,
                language,
                numbering_style,
            )
            target_words = int(sec.get("target_words") or 0) or self._section_word_target(strategy, total, idx)
            if is_agreement:
                activity_msg = (f"正在起草第 {idx+1}/{total} 条：{sec_title}" if language == "zh" else f"Drafting clause {idx+1}/{total}: {sec_title}")
            else:
                # sec_title already carries the section number (via
                # format_display_title which returns "{number} {clean}"); the
                # earlier "{sec_no} {sec_title}" form double-printed it as
                # "1 1 核心架构".
                activity_msg = (f"正在撰写第 {idx+1}/{total} 章：{sec_title}" if language == "zh" else f"Writing section {idx+1}/{total}: {sec_title}")
            yield {
                "type": "activity",
                "content": {
                    "kind": "writing",
                    "message": activity_msg,
                    "section": {"index": idx + 1, "total": total, "number": sec_no, "title": sec_title},
                },
            }
            section_md = await self._writer.write_section(
                title=sec_title,
                objective=str(sec.get("objective") or ""),
                key_points=[str(x) for x in (sec.get("key_points") or [])],
                evidence_focus=[str(x) for x in (sec.get("evidence_focus") or [])],
                must_cover_facts=self._normalize_section_facts(sec.get("must_cover_facts")),
                allowed_claim_types=[str(x) for x in (sec.get("allowed_claim_types") or [])],
                disallowed_claim_types=[str(x) for x in (sec.get("disallowed_claim_types") or [])],
                open_questions=[str(x) for x in (sec.get("open_questions") or [])],
                running_summary=self._compact_running_summary(running_summary),
                language=language,
                target_words=target_words,
                style_md=selected_style_md,
                level=sec_level,
                number=sec_no,
                visual_hint="none",
                source_refs=source_refs,
                writer_evidence_packet=packet.model_dump(),
                section_role=str(sec.get("role") or ("clause" if is_agreement else "analysis")),
                contract_context={**dict(body_contract_context or {}), "argument_items": list(sec.get("argument_items") or [])},
                debug_task_id=str(user_id or "anonymous"),
                prompt_variant=prompt_variant,
                numbering_style=numbering_style,
            )
            section_md = await self._writer.calibrate_length(section_markdown=section_md, language=language, target_words=target_words)

            # Phase 2B: per-section evaluation + regen.
            # Run only when standards exist (flag was on AND generation
            # succeeded). At most ONE regen per section — no infinite loop.
            # Issues are surfaced as a feedback block prepended to the
            # section's `objective`, so write_section sees them as part of
            # its own instructions and the writer's full original context
            # (publish channel, profile preset, formatting rules, etc.) is
            # preserved end-to-end.
            if section_standards and section_md.strip():
                try:
                    _, finder = self._eval_components()
                    section_issues = await finder.find(
                        content=section_md,
                        standards=section_standards,
                        scope="section",
                    )
                    actionable = [
                        i for i in section_issues
                        if i.severity in {"critical", "major"}
                    ]
                    if actionable:
                        log_print(
                            "[writer_engine][per_section_eval] section %s/%s '%s' "
                            "verdict=needs_repair issues=%s — regenerating once"
                            % (idx + 1, total, sec_title[:40], issues_summary(actionable)),
                            flush=True,
                        )
                        feedback = issues_as_writer_feedback(actionable)
                        original_objective = str(sec.get("objective") or "")
                        amended_objective = (
                            f"{original_objective}\n\n[修订指引]\n{feedback}"
                            if original_objective
                            else f"[修订指引]\n{feedback}"
                        )
                        section_md = await self._writer.write_section(
                            title=sec_title,
                            objective=amended_objective,
                            key_points=[str(x) for x in (sec.get("key_points") or [])],
                            evidence_focus=[str(x) for x in (sec.get("evidence_focus") or [])],
                            must_cover_facts=self._normalize_section_facts(sec.get("must_cover_facts")),
                            allowed_claim_types=[str(x) for x in (sec.get("allowed_claim_types") or [])],
                            disallowed_claim_types=[str(x) for x in (sec.get("disallowed_claim_types") or [])],
                            open_questions=[str(x) for x in (sec.get("open_questions") or [])],
                            running_summary=self._compact_running_summary(running_summary),
                            language=language,
                            target_words=target_words,
                            style_md=selected_style_md,
                            level=sec_level,
                            number=sec_no,
                            visual_hint="none",
                            source_refs=source_refs,
                            writer_evidence_packet=packet.model_dump(),
                            section_role=str(sec.get("role") or ("clause" if is_agreement else "analysis")),
                            contract_context={**dict(body_contract_context or {}), "argument_items": list(sec.get("argument_items") or [])},
                            debug_task_id=str(user_id or "anonymous"),
                            prompt_variant=prompt_variant,
                            numbering_style=numbering_style,
                        )
                        section_md = await self._writer.calibrate_length(
                            section_markdown=section_md,
                            language=language,
                            target_words=target_words,
                        )
                    else:
                        if section_issues:
                            log_print(
                                "[writer_engine][per_section_eval] section %s/%s minor-only issues=%d — keeping as-is"
                                % (idx + 1, total, len(section_issues)),
                                flush=True,
                            )
                except Exception as exc:
                    # Per-section eval is a quality enhancement, not a
                    # correctness gate — never let it break the run.
                    log_print(
                        f"[writer_engine][per_section_eval] section {idx+1}/{total} eval failed: {exc}",
                        flush=True,
                    )

            if not skip_visuals:
                sec_visual_hint = str(sec.get("visual_hint") or "chart").lower()
                visual_decision = await self._visual.decide_type(
                    section_title=sec_title,
                    section_markdown=section_md,
                    visual_hint=sec_visual_hint,
                    language=language,
                    infographic_budget_left=max_visuals - visuals_used,
                    contract_context=contract_context,
                )
                try:
                    log_print(
                        "[writer_engine] visual decision | section=%s type=%s confidence=%.2f reason=%s"
                        % (
                            sec_no,
                            str(visual_decision.get("type") or ""),
                            float(visual_decision.get("confidence") or 0.0),
                            str(visual_decision.get("reason") or "")[:160],
                        ),
                        flush=True,
                    )
                except Exception:
                    pass

                need_force = (
                    min_visuals > 0
                    and visuals_used < min_visuals
                    and (min_visuals - visuals_used) >= (total - idx)
                )
                if (
                    visuals_used < max_visuals
                    and (
                        need_force
                        or (
                            str(visual_decision.get("type") or "").strip().lower() == "infographic"
                            and float(visual_decision.get("confidence") or 0.0) >= 0.75
                        )
                    )
                ):
                    infographic = await self._maybe_generate_section_infographic(
                        section_no=sec_no,
                        section_title=sec_title,
                        section_markdown=section_md,
                        language=language,
                        user_id=user_id,
                    )
                    if infographic.get("ok"):
                        visuals_used += 1
                        img_url = str(infographic.get("url") or "").strip()
                        caption = str(infographic.get("caption") or ("信息图" if language == "zh" else "Infographic")).strip()
                        if img_url:
                            generated_urls.append(img_url)
                            section_md = section_md.rstrip() + f"\n\n![{caption}]({img_url})\n"
                if generated_urls:
                    section_md, _ = self._fill_non_http_image_links(section_md, generated_urls)

            section_md = MarkdownPostProcessor.sanitize_markdown_noise(section_md)
            section_md = MarkdownPostProcessor.normalize_section_headings(section_md, sec_level, language=language)
            section_md = MarkdownPostProcessor.normalize_list_item_style(section_md)
            section_md = MarkdownPostProcessor.ensure_section_citation(section_md, source_refs, language)

            running_summary += "\n- " + f"{sec_no} {sec_title}"
            total_len += len(section_md)
            yield {"type": "answer", "content": section_md.strip() + "\n\n"}

        required_images = min_visuals
        current_real = 0
        if required_images > 0:
            current_real = len([u for u in generated_urls if is_valid_remote_image_url(str(u).strip())])
            missing = required_images - current_real
            if missing > 0:
                self._log_stage(
                    "visual_minimum_unmet",
                    {
                        "required_visuals": required_images,
                        "actual_visuals": current_real,
                        "missing_visuals": missing,
                    },
                )

        refs = self._build_references_markdown(source_refs, language)
        if refs:
            yield {"type": "answer", "content": refs + "\n"}
        self._log_stage("structured_report_metrics", {"sections": total, "total_len": total_len})

    @staticmethod
    def _normalize_section_facts(value: Any) -> List[Dict[str, Any]]:
        items = value if isinstance(value, list) else [value] if value else []
        out: List[Dict[str, Any]] = []
        seen = set()
        for idx, item in enumerate(items, start=1):
            if isinstance(item, dict):
                payload = {
                    "fact_id": str(item.get("fact_id") or f"fact_{idx}").strip(),
                    "fact_type": str(item.get("fact_type") or "fact").strip(),
                    "summary": str(item.get("summary") or "").strip(),
                    "source_kind": str(item.get("source_kind") or "").strip(),
                    "source_ref": str(item.get("source_ref") or "").strip(),
                    "raw_evidence": str(item.get("raw_evidence") or item.get("summary") or "").strip(),
                }
            else:
                summary = str(item or "").strip()
                if not summary:
                    continue
                payload = {
                    "fact_id": f"fact_{idx}",
                    "fact_type": "fact",
                    "summary": summary,
                    "source_kind": "legacy_text",
                    "source_ref": "",
                    "raw_evidence": summary,
                }
            summary = str(payload.get("summary") or "").strip()
            if not summary or summary in seen:
                continue
            seen.add(summary)
            out.append(payload)
        return out

    # ── Translation pipeline ────────────────────────────────────────────

    TRANSLATABLE_EXTENSIONS = {".pdf", ".docx", ".xlsx", ".txt", ".md"}

    @staticmethod
    def _is_translatable_file(filename: str) -> bool:
        import os
        ext = os.path.splitext(str(filename or "").strip().lower())[1]
        return ext in WriterEnginePipeline.TRANSLATABLE_EXTENSIONS

    MAX_TRANSLATION_SEGMENT_CHARS = 6000

    @staticmethod
    def _split_long_translation_block(block: str, *, limit: int) -> List[str]:
        text = str(block or "")
        if len(text.strip()) <= limit:
            return [text] if text else []

        parts = re.split(r"(\n\s*\n+)", text)
        chunks: List[str] = []
        current = ""
        for part in parts:
            if not part:
                continue
            if current and len(current) + len(part) > limit:
                chunks.append(current)
                current = part
            else:
                current += part
        if current:
            chunks.append(current)

        out: List[str] = []
        sentence_boundary = re.compile(r"(?<=[。！？.!?])(\s+)")
        for chunk in chunks:
            if len(chunk.strip()) <= limit:
                out.append(chunk)
                continue
            sentences = sentence_boundary.split(chunk)
            current = ""
            for sentence in sentences:
                if current and len(current) + len(sentence) > limit:
                    out.append(current)
                    current = sentence
                else:
                    current += sentence
            if current:
                out.append(current)

        final: List[str] = []
        for chunk in out:
            if len(chunk.strip()) <= limit:
                final.append(chunk)
                continue
            for start in range(0, len(chunk), limit):
                final.append(chunk[start:start + limit])
        return [item for item in final if item.strip()]

    @classmethod
    def _split_translation_segments(cls, markdown: str) -> List[str]:
        """Split markdown for translation without cutting paragraphs first."""
        text = str(markdown or "").strip()
        if not text:
            return []
        segments: List[str] = []
        current: List[str] = []
        in_fence = False
        in_table = False
        limit = cls.MAX_TRANSLATION_SEGMENT_CHARS

        def flush_current() -> None:
            nonlocal current
            if current:
                segments.append("".join(current))
                current = []

        for line in text.splitlines(keepends=True):
            stripped = line.strip()
            if stripped.startswith("```"):
                in_fence = not in_fence
            is_table_line = stripped.startswith("|") and stripped.endswith("|")
            heading_boundary = bool(re.match(r"^#{1,6}\s+\S", stripped))
            page_boundary = bool(re.match(r"^(?:page|第)\s*\d+\b", stripped, flags=re.IGNORECASE))
            block_boundary = not stripped and not in_fence and not in_table

            should_flush = bool(current) and not in_fence and (
                heading_boundary
                or page_boundary
                or (block_boundary and len("".join(current)) >= limit)
            )
            if should_flush:
                flush_current()
                if block_boundary:
                    continue

            if is_table_line and not in_table and current and len("".join(current)) >= limit:
                flush_current()
            if not is_table_line and in_table and current and len("".join(current)) >= limit:
                flush_current()
            in_table = is_table_line
            current.append(line)
            if not in_fence and not in_table and len("".join(current)) >= limit * 2:
                flush_current()
        flush_current()

        bounded: List[str] = []
        for seg in segments:
            bounded.extend(cls._split_long_translation_block(seg, limit=limit))

        merged: List[str] = []
        for seg in bounded:
            if merged and len(seg.strip()) < 120 and len(merged[-1]) + len(seg) <= limit:
                merged[-1] += seg
            else:
                merged.append(seg)
        return merged or [text]

    @staticmethod
    def _extract_explicit_target_language(user_query: str) -> str:
        query = str(user_query or "").strip()
        if not query:
            return ""

        language_aliases = [
            ("Chinese", r"(?:中文|汉语|简体中文|繁体中文|Chinese|Mandarin)"),
            ("English", r"(?:英文|英语|English)"),
            ("Japanese", r"(?:日文|日语|Japanese)"),
            ("Korean", r"(?:韩文|韩语|朝鲜语|Korean)"),
            ("French", r"(?:法文|法语|French)"),
            ("German", r"(?:德文|德语|German)"),
            ("Spanish", r"(?:西班牙文|西班牙语|Spanish)"),
            ("Portuguese", r"(?:葡萄牙文|葡萄牙语|Portuguese)"),
            ("Russian", r"(?:俄文|俄语|Russian)"),
            ("Arabic", r"(?:阿拉伯文|阿拉伯语|Arabic)"),
        ]
        target_patterns = (
            r"(?:翻译|译|转译|translate|translated|translation)"
            r"(?:成|为|至|到|into|to)?\s*(?:目标语言(?:是|为)?\s*)?{alias}",
            r"(?:输出|生成|改写|转换)(?:成|为|至|到|into|to)?\s*{alias}(?:版|版本|文本|文档)?",
            r"(?:in|into|to)\s+{alias}\b",
        )
        for language, alias in language_aliases:
            for pattern in target_patterns:
                if re.search(pattern.format(alias=alias), query, re.I):
                    return language
        return ""

    async def _detect_target_language(self, user_query: str, source_language: str = "") -> str:
        explicit_target = self._extract_explicit_target_language(user_query)
        if explicit_target:
            return explicit_target

        from app.llm.types import Message, Role
        system = (
            "You extract the target language from a user's translation request.\n"
            "Return strict JSON only: {\"target_language\":\"English|Chinese|Japanese|Korean|French|German|Spanish|...\"}\n"
            "The source document language is: " + (source_language or "unknown") + ". "
            "The target language MUST be different from the source language.\n"
            "If the user does not specify a target language explicitly, "
            "default to English if the source is non-English, or Chinese if the source is English."
        )
        try:
            resp = await self._llm.ainvoke([
                Message(role=Role.SYSTEM, content=system),
                Message(role=Role.USER, content=user_query[:500]),
            ])
            raw = str(getattr(resp, "content", "") or "").strip()
            data = json.loads(raw.replace("```json", "").replace("```", "").strip())
            target = str(data.get("target_language") or "").strip()
            # This fallback only applies when the user did not explicitly name
            # a target language. Explicit user intent is resolved above.
            if target.lower() == source_language.lower():
                target = "English" if source_language != "English" else "Chinese"
            return target or "English"
        except Exception:
            return "English" if source_language != "English" else "Chinese"

    @staticmethod
    def _detect_source_language(text: str) -> str:
        sample = str(text or "")[:4000]
        sample = re.sub(
            r"[【\[]\s*(?:内嵌图片|嵌入图片|Embedded\s+Image)[^】\]]*[】\]]",
            " ",
            sample,
            flags=re.I,
        )
        chinese_count = len(re.findall(r"[\u4e00-\u9fff]", sample))
        latin_count = len(re.findall(r"[A-Za-z]", sample))
        if chinese_count >= 2 and (latin_count == 0 or chinese_count > latin_count * 0.15):
            return "Chinese"
        if re.search(r"[\u3040-\u309f\u30a0-\u30ff]", sample):
            return "Japanese"
        if re.search(r"[\uac00-\ud7af]", sample):
            return "Korean"
        return "English"

    async def generate_translation(
        self,
        *,
        user_query: str,
        language: str,
        evidence_bundle: Dict[str, Any],
        user_id: str = "anonymous",
    ) -> AsyncIterator[Dict[str, Any]]:
        from app.utils.i18n import get as i18n_get

        source_markdown = str((evidence_bundle or {}).get("translation_source_markdown") or "").strip()
        source_filename = str((evidence_bundle or {}).get("translation_source_filename") or "").strip()
        source_ext = str((evidence_bundle or {}).get("translation_source_ext") or "").strip().lower()
        source_object_path = str((evidence_bundle or {}).get("translation_source_object_path") or "").strip()
        source_signed_url = str((evidence_bundle or {}).get("translation_source_url") or "").strip()

        # Validate: must have uploaded document (markdown 或原始二进制其一即可)
        if not source_markdown and not source_object_path and not source_signed_url:
            yield {"type": "answer", "content": i18n_get("translation_no_document", language) + "\n"}
            return

        # Validate: file format
        if source_filename and not self._is_translatable_file(source_filename):
            yield {"type": "answer", "content": i18n_get("translation_unsupported_format", language) + "\n"}
            return

        # .docx 走原位翻译（保留样式、合并单元格、页眉页脚、表格结构）
        if source_ext == ".docx" and (source_object_path or source_signed_url):
            async for evt in self._generate_translation_docx_inplace(
                user_query=user_query,
                language=language,
                source_filename=source_filename,
                source_object_path=source_object_path,
                source_signed_url=source_signed_url,
                user_id=user_id,
            ):
                yield evt
            return

        # .xlsx 走原位翻译（保留 workbook 结构、样式、公式、数字/日期）
        if source_ext == ".xlsx" and (source_object_path or source_signed_url):
            async for evt in self._generate_translation_xlsx_inplace(
                user_query=user_query,
                language=language,
                source_filename=source_filename,
                source_object_path=source_object_path,
                source_signed_url=source_signed_url,
                user_id=user_id,
            ):
                yield evt
            return

        # Detect languages
        yield {
            "type": "activity",
            "content": {
                "kind": "writing",
                "message": i18n_get("translation_detecting_language", language),
            },
        }
        source_lang = self._detect_source_language(source_markdown)
        target_lang = await self._detect_target_language(user_query, source_language=source_lang)

        # Build translation prompt
        translator_prompt = self._prompts.get(
            "translator",
            source_language=source_lang,
            target_language=target_lang,
        )

        # Split into segments
        segments = self._split_translation_segments(source_markdown)
        total = len(segments)

        self._log_stage("translation_start", {
            "source_lang": source_lang,
            "target_lang": target_lang,
            "source_chars": len(source_markdown),
            "segments": total,
            "filename": source_filename,
        })

        # Translate each segment
        from app.llm.types import Message, Role
        for idx, segment in enumerate(segments):
            yield {
                "type": "activity",
                "content": {
                    "kind": "writing",
                    "message": i18n_get("translation_translating_segment", language).format(
                        current=idx + 1, total=total
                    ),
                },
            }
            try:
                resp = await self._llm.ainvoke([
                    Message(role=Role.SYSTEM, content=translator_prompt),
                    Message(role=Role.USER, content=segment),
                ])
                translated = str(getattr(resp, "content", "") or "").strip()
                if translated:
                    yield {"type": "answer", "content": translated + "\n\n"}
                else:
                    # Fallback: output original if translation fails
                    yield {"type": "answer", "content": segment + "\n\n"}
            except Exception as e:
                self._log_stage("translation_segment_error", {
                    "segment": idx + 1, "error": str(e)[:260]
                })
                yield {"type": "answer", "content": segment + "\n\n"}

        self._log_stage("translation_complete", {
            "source_lang": source_lang,
            "target_lang": target_lang,
            "segments": total,
        })

    async def _generate_translation_docx_inplace(
        self,
        *,
        user_query: str,
        language: str,
        source_filename: str,
        source_object_path: str,
        source_signed_url: str,
        user_id: str,
    ) -> AsyncIterator[Dict[str, Any]]:
        """In-place .docx translation.
        下载原始 .docx → 原位翻译（保留样式/合并单元格/页眉页脚） → 上传 OSS → 发布 document event。
        """
        import os as _os
        from app.utils.i18n import get as i18n_get
        from app.utils.oss_uploader import AliyunOSSUploader
        from app.services.translation.docx_inplace import extract_segments, translate_docx_inplace

        # 1) 获取原始 .docx 二进制
        yield {
            "type": "activity",
            "content": {
                "kind": "writing",
                "message": i18n_get("translation_detecting_language", language),
            },
        }
        uploader = AliyunOSSUploader()
        download_url = source_signed_url
        source_bytes = b""
        if source_object_path:
            try:
                source_bytes = uploader.read_bytes(source_object_path)
            except Exception:
                try:
                    download_url = uploader.internal_url(source_object_path)
                except Exception:
                    download_url = source_signed_url
        if not source_bytes and not download_url:
            yield {"type": "answer", "content": i18n_get("translation_no_document", language) + "\n"}
            return

        import httpx
        if not source_bytes:
            try:
                async with httpx.AsyncClient(timeout=60.0) as client:
                    resp = await client.get(download_url)
                    resp.raise_for_status()
                    source_bytes = resp.content
            except Exception as exc:
                self._log_stage("translation_docx_fetch_failed", {"error": str(exc)[:240]})
                yield {"type": "answer", "content": i18n_get("translation_no_document", language) + "\n"}
                return

        # 2) 推断源/目标语言（先看文件内容少量字节，解不准就用中文兜底）
        try:
            from docx import Document as _Doc
            import io as _io
            _probe = _Doc(_io.BytesIO(source_bytes))
            _sample = "\n".join(seg.text for seg in extract_segments(_probe)[:30])[:1000]
        except Exception:
            _sample = ""
        source_lang = self._detect_source_language(_sample or source_filename or "")
        target_lang = await self._detect_target_language(user_query, source_language=source_lang)

        self._log_stage("translation_inplace_start", {
            "filename": source_filename,
            "source_lang": source_lang,
            "target_lang": target_lang,
            "bytes": len(source_bytes),
        })

        yield {
            "type": "activity",
            "content": {
                "kind": "writing",
                "message": i18n_get("translation_translating_segment", language).format(current=1, total=1),
            },
        }

        # 3) 执行原位翻译
        try:
            result = await translate_docx_inplace(
                source_bytes=source_bytes,
                source_language=source_lang,
                target_language=target_lang,
                llm=self._llm,
            )
        except Exception as exc:
            self._log_stage("translation_inplace_failed", {"error": str(exc)[:240]})
            yield {"type": "answer", "content": f"{i18n_get('translation_complete', language)} (fallback: see markdown output)\n"}
            return

        # 4) 上传到 OSS
        base = _os.path.splitext(source_filename or "translated_document")[0]
        out_filename = f"{base}_{target_lang}.docx"
        try:
            signed_url, object_path = uploader.upload_bytes_with_path(
                content=result.file_bytes,
                user_id=str(user_id or "anonymous"),
                file_name=out_filename,
                content_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            )
            final_signed = uploader.sign_url(object_path)
        except Exception as exc:
            self._log_stage("translation_inplace_upload_failed", {"error": str(exc)[:240]})
            yield {"type": "answer", "content": f"Upload failed: {exc}\n"}
            return

        self._log_stage("translation_inplace_complete", {
            **result.stats,
            "filename": out_filename,
        })

        # 5) 发布 document event（前端据此展示下载卡片）
        yield {
            "type": "document",
            "content": {
                "type": "docx",
                "url": final_signed,
                "filename": out_filename,
                "title": base,
                "object_path": object_path,
                "render_stats": result.stats,
            },
        }
        # 给个简短的 answer 作为完成标志，避免空 content 让上游卡死
        yield {"type": "answer", "content": f"{i18n_get('translation_complete', language)}\n"}

    async def _generate_translation_xlsx_inplace(
        self,
        *,
        user_query: str,
        language: str,
        source_filename: str,
        source_object_path: str,
        source_signed_url: str,
        user_id: str,
    ) -> AsyncIterator[Dict[str, Any]]:
        """In-place .xlsx translation.
        下载原始 .xlsx → 翻译文本单元格 → 上传 OSS → 发布 document event。
        """
        import os as _os
        from app.utils.i18n import get as i18n_get
        from app.utils.oss_uploader import AliyunOSSUploader
        from app.services.translation.xlsx_inplace import extract_xlsx_segments, translate_xlsx_inplace

        yield {
            "type": "activity",
            "content": {
                "kind": "writing",
                "message": i18n_get("translation_detecting_language", language),
            },
        }
        uploader = AliyunOSSUploader()
        download_url = source_signed_url
        source_bytes = b""
        if source_object_path:
            try:
                source_bytes = uploader.read_bytes(source_object_path)
            except Exception:
                try:
                    download_url = uploader.internal_url(source_object_path)
                except Exception:
                    download_url = source_signed_url
        if not source_bytes and not download_url:
            yield {"type": "answer", "content": i18n_get("translation_no_document", language) + "\n"}
            return

        import httpx
        if not source_bytes:
            try:
                async with httpx.AsyncClient(timeout=60.0) as client:
                    resp = await client.get(download_url)
                    resp.raise_for_status()
                    source_bytes = resp.content
            except Exception as exc:
                self._log_stage("translation_xlsx_fetch_failed", {"error": str(exc)[:240]})
                yield {"type": "answer", "content": i18n_get("translation_no_document", language) + "\n"}
                return

        try:
            from openpyxl import load_workbook as _load_workbook
            import io as _io
            _probe = _load_workbook(_io.BytesIO(source_bytes), data_only=False, read_only=True)
            _sample = "\n".join(seg.text for seg in extract_xlsx_segments(_probe)[:80])[:1000]
            try:
                _probe.close()
            except Exception:
                pass
        except Exception:
            _sample = ""
        source_lang = self._detect_source_language(_sample or source_filename or "")
        target_lang = await self._detect_target_language(user_query, source_language=source_lang)

        self._log_stage("translation_xlsx_start", {
            "filename": source_filename,
            "source_lang": source_lang,
            "target_lang": target_lang,
            "bytes": len(source_bytes),
        })

        yield {
            "type": "activity",
            "content": {
                "kind": "writing",
                "message": i18n_get("translation_translating_spreadsheet", language),
            },
        }

        try:
            result = await translate_xlsx_inplace(
                source_bytes=source_bytes,
                source_language=source_lang,
                target_language=target_lang,
                llm=self._llm,
            )
        except Exception as exc:
            self._log_stage("translation_xlsx_failed", {"error": str(exc)[:240]})
            yield {"type": "answer", "content": f"{i18n_get('translation_unsupported_format', language)}\n"}
            return

        base = _os.path.splitext(source_filename or "translated_spreadsheet")[0]
        out_filename = f"{base}_{target_lang}.xlsx"
        try:
            _signed_url, object_path = uploader.upload_bytes_with_path(
                content=result.file_bytes,
                user_id=str(user_id or "anonymous"),
                file_name=out_filename,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            final_signed = uploader.sign_url(object_path)
        except Exception as exc:
            self._log_stage("translation_xlsx_upload_failed", {"error": str(exc)[:240]})
            yield {"type": "answer", "content": f"Upload failed: {exc}\n"}
            return

        self._log_stage("translation_xlsx_complete", {
            **result.stats,
            "filename": out_filename,
        })

        yield {
            "type": "document",
            "content": {
                "type": "xlsx",
                "url": final_signed,
                "filename": out_filename,
                "title": base,
                "object_path": object_path,
                "render_stats": result.stats,
            },
        }
        yield {"type": "answer", "content": f"{i18n_get('translation_complete', language)}\n"}
