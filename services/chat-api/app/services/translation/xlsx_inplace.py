"""In-place XLSX translator.

The translator preserves workbook structure, formulas, numeric/date values,
cell styles, merged cells, filters, charts, and drawings as much as openpyxl
allows. It only replaces plain text cell values.
"""

from __future__ import annotations

import io
import json
import logging
import re
from dataclasses import dataclass, field
from typing import Any, Dict, List

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

from app.llm.factory import get_llm_client
from app.llm.types import Message, Role


logger = logging.getLogger(__name__)


@dataclass
class XlsxSegment:
    seg_id: int
    text: str
    sheet_name: str
    coordinate: str
    cell_ref: Cell = field(repr=False)
    row_context: List[str] = field(default_factory=list)
    column_header: str = ""


@dataclass
class XlsxTranslationResult:
    file_bytes: bytes
    stats: Dict[str, Any]
    glossary: Dict[str, str]


_PLACEHOLDER_ONLY = re.compile(r"^\s*[\d\s\.\,\-\/\(\)\[\]\{\}#@%\*\+\$¥￥:：]*\s*$")
_INVALID_SHEET_TITLE_CHARS = re.compile(r"[\\/*?:\[\]]+")


def _text_is_translatable(text: str) -> bool:
    value = str(text or "").strip()
    if not value:
        return False
    if value.startswith("="):
        return False
    if _PLACEHOLDER_ONLY.match(value):
        return False
    return True


def _string_cell_value(cell: Cell) -> str:
    value = cell.value
    if not isinstance(value, str):
        return ""
    if cell.data_type == "f":
        return ""
    return value.strip()


def _row_context(ws: Any, row_idx: int, col_idx: int) -> List[str]:
    values: List[str] = []
    try:
        for cell in ws[row_idx]:
            if cell.column == col_idx:
                continue
            text = _string_cell_value(cell)
            if text and len(text) <= 120:
                values.append(text)
    except Exception:
        return []
    return values[:8]


def _column_header(ws: Any, col_idx: int, row_idx: int) -> str:
    if row_idx <= 1:
        return ""
    try:
        text = _string_cell_value(ws.cell(row=1, column=col_idx))
        return text[:120]
    except Exception:
        return ""


def extract_xlsx_segments(workbook: Any) -> List[XlsxSegment]:
    segments: List[XlsxSegment] = []
    next_id = 0
    for ws in workbook.worksheets:
        if ws.sheet_state != "visible":
            continue
        for row in ws.iter_rows():
            for cell in row:
                text = _string_cell_value(cell)
                if not _text_is_translatable(text):
                    continue
                segments.append(
                    XlsxSegment(
                        seg_id=next_id,
                        text=text,
                        sheet_name=str(ws.title or ""),
                        coordinate=str(cell.coordinate),
                        cell_ref=cell,
                        row_context=_row_context(ws, int(cell.row), int(cell.column)),
                        column_header=_column_header(ws, int(cell.column), int(cell.row)),
                    )
                )
                next_id += 1
    return segments


async def extract_glossary(
    *,
    llm: Any,
    full_text_sample: str,
    source_language: str,
    target_language: str,
) -> Dict[str, str]:
    system = (
        "You extract a minimal glossary for spreadsheet translation.\n"
        f"Translate from {source_language} to {target_language}.\n"
        "Include recurring business terms, proper nouns, product names, department names, "
        "status labels, and domain-specific terms.\n"
        "EXCLUDE generic words and one-off text.\n"
        'Return strict JSON: {"glossary": [{"source": "...", "target": "..."}, ...]}\n'
        "Keep it under 40 entries."
    )
    payload = {"spreadsheet_excerpt": full_text_sample[:8000]}
    try:
        resp = await llm.ainvoke(
            [
                Message(role=Role.SYSTEM, content=system),
                Message(role=Role.USER, content=json.dumps(payload, ensure_ascii=False)),
            ]
        )
        raw = str(getattr(resp, "content", "") or "").strip()
        data = json.loads(raw.replace("```json", "").replace("```", "").strip())
        out: Dict[str, str] = {}
        for item in (data or {}).get("glossary") or []:
            if not isinstance(item, dict):
                continue
            src = str(item.get("source") or "").strip()
            tgt = str(item.get("target") or "").strip()
            if src and tgt and src != tgt:
                out[src] = tgt
        return out
    except Exception as exc:
        logger.warning("[xlsx_inplace] glossary_extraction_failed: %s", exc)
        return {}


_TRANSLATION_SYSTEM = """You are a professional spreadsheet translator.
Translate from {source_language} to {target_language}.

Rules:
- Preserve ALL numbers, dates, currencies, units, placeholders, codes, IDs, URLs, and email addresses verbatim
- Keep spreadsheet cell text concise
- Keep proper nouns, brand names, and technical terms as-is unless the glossary says otherwise
- Use the provided glossary strictly for listed terms
- Translate EACH cell independently; do not merge, split, or reorder cells
- Do NOT add translator notes, explanations, bullets, or extra punctuation
- Output strict JSON with the same ids
"""


async def translate_batch(
    *,
    llm: Any,
    batch: List[XlsxSegment],
    source_language: str,
    target_language: str,
    glossary: Dict[str, str],
    context_before: str = "",
    context_after: str = "",
) -> Dict[int, str]:
    if not batch:
        return {}
    system = _TRANSLATION_SYSTEM.format(
        source_language=source_language,
        target_language=target_language,
    )
    payload: Dict[str, Any] = {
        "glossary": [{"source": k, "target": v} for k, v in glossary.items()][:40],
        "context_before": context_before[:400],
        "context_after": context_after[:400],
        "cells": [
            {
                "id": seg.seg_id,
                "sheet": seg.sheet_name,
                "cell": seg.coordinate,
                "text": seg.text,
                **({"row_context": seg.row_context} if seg.row_context else {}),
                **({"column_header": seg.column_header} if seg.column_header else {}),
            }
            for seg in batch
        ],
    }
    try:
        resp = await llm.ainvoke(
            [
                Message(role=Role.SYSTEM, content=system),
                Message(
                    role=Role.USER,
                    content=(
                        "Translate each spreadsheet cell. Return strict JSON:\n"
                        '{"translations":[{"id":<int>,"text":"<translated>"}]}\n\n'
                        + json.dumps(payload, ensure_ascii=False)
                    ),
                ),
            ]
        )
        raw = str(getattr(resp, "content", "") or "").strip()
        data = json.loads(raw.replace("```json", "").replace("```", "").strip())
        out: Dict[int, str] = {}
        for item in (data or {}).get("translations") or []:
            if not isinstance(item, dict):
                continue
            try:
                sid = int(item.get("id"))
            except Exception:
                continue
            text = str(item.get("text") or "").strip()
            if text:
                out[sid] = text
        return out
    except Exception as exc:
        logger.warning("[xlsx_inplace] batch_translation_failed: %s", exc)
        return {}


def _safe_sheet_title(value: str, existing: set[str], fallback: str) -> str:
    title = _INVALID_SHEET_TITLE_CHARS.sub("_", str(value or "").strip())
    title = title.replace("\n", " ").replace("\r", " ").strip("' ")
    if not title:
        title = fallback
    title = title[:31] or fallback[:31] or "Sheet"
    if title not in existing:
        return title
    base = title[:28].rstrip() or "Sheet"
    idx = 2
    while True:
        candidate = f"{base}_{idx}"[:31]
        if candidate not in existing:
            return candidate
        idx += 1


async def translate_sheet_titles(
    *,
    llm: Any,
    sheet_names: List[str],
    source_language: str,
    target_language: str,
    glossary: Dict[str, str],
) -> Dict[str, str]:
    translatable = [name for name in sheet_names if _text_is_translatable(name)]
    if not translatable:
        return {}
    system = (
        "You translate Excel worksheet tab names.\n"
        f"Translate from {source_language} to {target_language}.\n"
        "Keep names concise, natural, and suitable for Excel sheet tabs.\n"
        "Do not add explanations. Preserve IDs/codes when present.\n"
        "Return strict JSON with the same ids."
    )
    payload = {
        "glossary": [{"source": k, "target": v} for k, v in glossary.items()][:40],
        "sheet_names": [{"id": idx, "text": name} for idx, name in enumerate(translatable)],
    }
    try:
        resp = await llm.ainvoke(
            [
                Message(role=Role.SYSTEM, content=system),
                Message(
                    role=Role.USER,
                    content=(
                        "Translate each sheet name. Return strict JSON:\n"
                        '{"translations":[{"id":<int>,"text":"<translated>"}]}\n\n'
                        + json.dumps(payload, ensure_ascii=False)
                    ),
                ),
            ]
        )
        raw = str(getattr(resp, "content", "") or "").strip()
        data = json.loads(raw.replace("```json", "").replace("```", "").strip())
        out: Dict[str, str] = {}
        for item in (data or {}).get("translations") or []:
            if not isinstance(item, dict):
                continue
            try:
                idx = int(item.get("id"))
            except Exception:
                continue
            if idx < 0 or idx >= len(translatable):
                continue
            text = str(item.get("text") or "").strip()
            if text:
                out[translatable[idx]] = text
        return out
    except Exception as exc:
        logger.warning("[xlsx_inplace] sheet_title_translation_failed: %s", exc)
        return {}
async def translate_xlsx_inplace(
    *,
    source_bytes: bytes,
    source_language: str,
    target_language: str,
    llm: Any = None,
    batch_size: int = 40,
) -> XlsxTranslationResult:
    if llm is None:
        llm = get_llm_client(streaming=False, stage="compose")

    wb = load_workbook(io.BytesIO(source_bytes), data_only=False)
    segments = extract_xlsx_segments(wb)
    if not segments:
        buf = io.BytesIO()
        wb.save(buf)
        return XlsxTranslationResult(
            file_bytes=buf.getvalue(),
            stats={"segments": 0, "segments_translated": 0, "batches": 0, "sheets": len(wb.worksheets)},
            glossary={},
        )

    full_sample = "\n".join(seg.text for seg in segments)[:8000]
    glossary = await extract_glossary(
        llm=llm,
        full_text_sample=full_sample,
        source_language=source_language,
        target_language=target_language,
    )

    translations: Dict[int, str] = {}
    total = len(segments)
    num_batches = (total + batch_size - 1) // batch_size
    for b_idx in range(num_batches):
        start = b_idx * batch_size
        end = min(start + batch_size, total)
        context_before = segments[start - 1].text if start > 0 else ""
        context_after = segments[end].text if end < total else ""
        translations.update(
            await translate_batch(
                llm=llm,
                batch=segments[start:end],
                source_language=source_language,
                target_language=target_language,
                glossary=glossary,
                context_before=context_before,
                context_after=context_after,
            )
        )

    written = 0
    for seg in segments:
        translated = translations.get(seg.seg_id)
        if not translated:
            continue
        seg.cell_ref.value = translated
        written += 1

    original_sheet_names = [str(ws.title or "") for ws in wb.worksheets]
    sheet_title_map = await translate_sheet_titles(
        llm=llm,
        sheet_names=original_sheet_names,
        source_language=source_language,
        target_language=target_language,
        glossary=glossary,
    )
    renamed = 0
    existing_titles: set[str] = set()
    for ws in wb.worksheets:
        original_title = str(ws.title or "")
        translated_title = sheet_title_map.get(original_title, "")
        if not translated_title or translated_title == original_title:
            existing_titles.add(original_title)
            continue
        new_title = _safe_sheet_title(translated_title, existing_titles, original_title)
        if new_title != original_title:
            ws.title = new_title
            renamed += 1
        existing_titles.add(str(ws.title or ""))

    buf = io.BytesIO()
    wb.save(buf)
    return XlsxTranslationResult(
        file_bytes=buf.getvalue(),
        stats={
            "segments": total,
            "segments_translated": written,
            "batches": num_batches,
            "sheets": len(wb.worksheets),
            "sheet_titles_translated": renamed,
            "glossary_size": len(glossary),
        },
        glossary=glossary,
    )


__all__ = [
    "XlsxSegment",
    "XlsxTranslationResult",
    "extract_xlsx_segments",
    "translate_batch",
    "translate_sheet_titles",
    "translate_xlsx_inplace",
]
