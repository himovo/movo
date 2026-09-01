from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from app.enterprise_capabilities.spreadsheets.models import SpreadsheetSpec


class SpreadsheetValidationError(ValueError):
    """Raised when rendered XLSX bytes do not match their source specification."""


def validate_spreadsheet_spec(spec: SpreadsheetSpec) -> None:
    if not spec.sheets:
        raise SpreadsheetValidationError("spreadsheet has no sheets")
    for sheet in spec.sheets:
        if not sheet.columns:
            raise SpreadsheetValidationError(f"sheet '{sheet.name}' has no columns")
        if not sheet.rows:
            raise SpreadsheetValidationError(f"sheet '{sheet.name}' has no data rows")


def validate_rendered_workbook(content: bytes, spec: SpreadsheetSpec) -> dict[str, Any]:
    validate_spreadsheet_spec(spec)
    try:
        workbook = load_workbook(BytesIO(content), read_only=False, data_only=False)
    except Exception as exc:
        raise SpreadsheetValidationError(f"rendered workbook cannot be opened: {exc}") from exc

    summaries: list[dict[str, Any]] = []
    try:
        expected_names = [sheet.name for sheet in spec.sheets]
        if workbook.sheetnames != expected_names:
            raise SpreadsheetValidationError(
                f"rendered sheet names do not match: expected {expected_names}, got {workbook.sheetnames}"
            )
        for sheet_spec in spec.sheets:
            worksheet = workbook[sheet_spec.name]
            expected_headers = [column.title for column in sheet_spec.columns]
            actual_headers = [worksheet.cell(1, index).value for index in range(1, len(expected_headers) + 1)]
            if actual_headers != expected_headers:
                raise SpreadsheetValidationError(
                    f"sheet '{sheet_spec.name}' headers do not match: expected {expected_headers}, got {actual_headers}"
                )
            expected_rows = len(sheet_spec.rows)
            nonempty_rows = sum(
                1
                for row in worksheet.iter_rows(min_row=2, values_only=True)
                if any(value is not None and str(value).strip() for value in row)
            )
            if nonempty_rows != expected_rows:
                raise SpreadsheetValidationError(
                    f"sheet '{sheet_spec.name}' row count does not match: expected {expected_rows}, got {nonempty_rows}"
                )
            summaries.append({
                "name": sheet_spec.name,
                "column_count": len(expected_headers),
                "data_row_count": nonempty_rows,
                "headers": expected_headers,
            })
    finally:
        workbook.close()
    return {"verified": True, "sheet_count": len(summaries), "sheets": summaries}
