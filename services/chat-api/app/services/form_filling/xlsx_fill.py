from __future__ import annotations

import io
import re
from copy import copy
from typing import Any, Dict, List, Tuple

from openpyxl.cell.cell import MergedCell
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

from app.services.form_filling.mapper import build_fill_plan
from app.services.form_filling.models import FillResult


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return ""


def _target(sheet: str, coord: str) -> str:
    return f"{sheet}!{coord}"


def _merged_anchor_coordinate(ws: Any, row: int, col: int) -> str:
    coord = ws.cell(row, col).coordinate
    for merged_range in list(ws.merged_cells.ranges):
        if coord in merged_range:
            return ws.cell(merged_range.min_row, merged_range.min_col).coordinate
    return coord


def _writable_cell(ws: Any, coord: str) -> Any:
    cell = ws[coord]
    if not isinstance(cell, MergedCell):
        return cell
    for merged_range in list(ws.merged_cells.ranges):
        if coord in merged_range:
            return ws.cell(merged_range.min_row, merged_range.min_col)
    return cell


def extract_xlsx_form_schema(source_bytes: bytes) -> Dict[str, Any]:
    wb = load_workbook(io.BytesIO(source_bytes), data_only=False)
    fill_targets: List[Dict[str, Any]] = []
    tables: List[Dict[str, Any]] = []
    sparse_resume_targets: List[Dict[str, Any]] = []
    try:
        for ws in wb.worksheets:
            if ws.sheet_state != "visible":
                continue
            max_row = min(ws.max_row or 1, 120)
            max_col = min(ws.max_column or 1, 40)
            for row in range(1, max_row + 1):
                text_cells: List[Tuple[int, str]] = []
                for col in range(1, max_col + 1):
                    text = _cell_text(ws.cell(row, col).value)
                    if text:
                        text_cells.append((col, text))
                        right = ws.cell(row, col + 1) if col + 1 <= max_col else None
                        below = ws.cell(row + 1, col) if row + 1 <= max_row else None
                        if right is not None and right.value in (None, ""):
                            target_coord = _merged_anchor_coordinate(ws, right.row, right.column)
                            fill_targets.append(
                                {
                                    "target": _target(ws.title, target_coord),
                                    "label": text[:80],
                                    "kind": "right_empty_cell",
                                }
                            )
                        if below is not None and below.value in (None, ""):
                            target_coord = _merged_anchor_coordinate(ws, below.row, below.column)
                            fill_targets.append(
                                {
                                    "target": _target(ws.title, target_coord),
                                    "label": text[:80],
                                    "kind": "below_empty_cell",
                                }
                            )
                if len(text_cells) >= 2:
                    headers = [text for _, text in text_cells]
                    tables.append(
                        {
                            "target": f"{ws.title}!table:{row}",
                            "sheet": ws.title,
                            "header_row": row,
                            "headers": headers[:30],
                        }
                    )
            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    text = _cell_text(ws.cell(row, col).value)
                    for name in re.findall(r"\{\{\s*([^{}]{1,50})\s*\}\}", text):
                        fill_targets.append(
                            {
                                "target": _target(ws.title, ws.cell(row, col).coordinate),
                                "label": name.strip(),
                                "kind": "placeholder_cell",
                            }
                        )
            sparse_resume_targets.extend(_sparse_resume_template_targets(ws))
    finally:
        try:
            wb.close()
        except Exception:
            pass
    if len(fill_targets) <= 2 and not tables and sparse_resume_targets:
        fill_targets = sparse_resume_targets
    return {"fill_targets": fill_targets[:300], "tables": tables[:80]}


def _non_empty_text_cells(ws: Any, max_row: int = 120, max_col: int = 40) -> List[str]:
    values: List[str] = []
    for row in range(1, min(ws.max_row or 1, max_row) + 1):
        for col in range(1, min(ws.max_column or 1, max_col) + 1):
            text = _cell_text(ws.cell(row, col).value)
            if text:
                values.append(text)
    return values


def _sparse_resume_template_targets(ws: Any) -> List[Dict[str, Any]]:
    """Fallback for visually formatted resume templates with almost no cell text.

    Some downloaded XLSX resume templates are blank grids plus an image/photo or
    WPS drawing. There are no labels for schema extraction, so we expose the
    prominent merged cells as semantic resume fields.
    """

    texts = _non_empty_text_cells(ws)
    if len(texts) > 2:
        return []
    title_text = " ".join(texts)
    has_resume_title = any(token in title_text for token in ("个人简历", "简历", "resume", "Resume"))
    merged_refs = {str(rng) for rng in list(ws.merged_cells.ranges)}
    looks_like_resume_grid = (ws.max_row or 0) >= 25 and (ws.max_column or 0) >= 7 and "A1:H2" in merged_refs
    if not (has_resume_title or looks_like_resume_grid):
        return []

    labels = [
        ("D6", "姓名、英文名和当前身份"),
        ("D7", "性别"),
        ("E7", "出生日期"),
        ("F7", "国籍/所在地"),
        ("D8", "联系方式或公开主页"),
        ("B10", "基本信息栏标题"),
        ("B11", "最高学历"),
        ("D11", "毕业院校/专业"),
        ("G11", "工作年限"),
        ("B12", "个人简介/求职概况"),
        ("B14", "求职意向"),
        ("D14", "核心职能"),
        ("G14", "可到岗/工作地点"),
        ("B15", "核心能力/技能标签"),
        ("B17", "教育背景栏标题"),
        ("B18", "教育经历一"),
        ("D18", "教育经历二"),
        ("G18", "教育补充"),
        ("B19", "证书/荣誉"),
        ("D19", "语言/技能补充"),
        ("B20", "工作经历栏标题"),
        ("B21", "工作经历一时间"),
        ("C21", "工作经历一公司和职位"),
        ("E21", "工作经历一职责/成果"),
        ("B22", "工作经历二时间"),
        ("C22", "工作经历二公司和职位"),
        ("E22", "工作经历二职责/成果"),
        ("B23", "项目经历/代表成果栏标题"),
        ("B24", "项目或成就一"),
        ("D24", "项目或成就二"),
        ("F24", "项目或成就三"),
        ("B25", "项目或成就四"),
        ("D25", "项目或成就五"),
        ("F25", "项目或成就六"),
        ("B27", "详细履历/补充说明"),
    ]
    targets: List[Dict[str, Any]] = []
    for coord, label in labels:
        try:
            cell = _writable_cell(ws, coord)
            if isinstance(cell, MergedCell):
                continue
            targets.append({"target": _target(ws.title, cell.coordinate), "label": label, "kind": "sparse_resume_template"})
        except Exception:
            continue
    return targets


def _split_xlsx_target(target: str) -> Tuple[str, str]:
    sheet, coord = str(target or "").split("!", 1)
    return sheet, coord


def _copy_row_style(ws: Any, source_row: int, target_row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)
        if isinstance(src, MergedCell) or isinstance(dst, MergedCell):
            continue
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.font:
            dst.font = copy(src.font)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.border:
            dst.border = copy(src.border)


def _style_filled_cell(cell: Any, value: str, kind: str = "") -> None:
    text = str(value or "")
    cell.alignment = copy(cell.alignment) if cell.alignment else Alignment()
    cell.alignment = Alignment(
        horizontal=cell.alignment.horizontal or "left",
        vertical=cell.alignment.vertical or "center",
        text_rotation=cell.alignment.text_rotation,
        wrap_text=True,
        shrink_to_fit=cell.alignment.shrink_to_fit,
        indent=cell.alignment.indent,
    )
    if kind == "sparse_resume_template":
        cell.font = copy(cell.font) if cell.font else Font()
        cell.font = Font(
            name=cell.font.name,
            sz=10 if len(text) <= 80 else 9,
            bold=cell.font.bold,
            italic=cell.font.italic,
            vertAlign=cell.font.vertAlign,
            underline=cell.font.underline,
            strike=cell.font.strike,
            color=cell.font.color,
        )


async def fill_xlsx_form(
    *,
    source_bytes: bytes,
    user_text: str,
    llm: Any,
    overwrite: bool = False,
) -> FillResult:
    schema = extract_xlsx_form_schema(source_bytes)
    plan = await build_fill_plan(llm=llm, user_text=user_text, schema=schema, file_type="xlsx", overwrite=overwrite)
    wb = load_workbook(io.BytesIO(source_bytes), data_only=False)
    warnings: List[str] = [str(w) for w in list(plan.get("warnings") or []) if str(w).strip()]
    filled = 0
    appended = 0
    skipped = 0
    cleared_placeholders = 0
    unresolved: List[str] = []
    try:
        target_kind = {
            str(t.get("target") or ""): str(t.get("kind") or "")
            for t in list(schema.get("fill_targets") or [])
            if isinstance(t, dict)
        }
        for item in list(plan.get("fills") or []):
            if not isinstance(item, dict):
                continue
            value = str(item.get("value") or "").strip()
            if not value:
                continue
            try:
                sheet, coord = _split_xlsx_target(str(item.get("target") or ""))
                ws = wb[sheet]
                cell = _writable_cell(ws, coord)
                if isinstance(cell, MergedCell):
                    skipped += 1
                    continue
            except Exception:
                skipped += 1
                continue
            old = cell.value
            if old not in (None, "") and not overwrite:
                text = str(old or "")
                if "{{" in text and "}}" in text:
                    label = str(item.get("label") or "").strip()
                    cell.value = re.sub(r"\{\{\s*" + re.escape(label) + r"\s*\}\}", value, text)
                    _style_filled_cell(cell, value, target_kind.get(str(item.get("target") or ""), ""))
                    filled += 1
                else:
                    skipped += 1
                continue
            cell.value = value
            _style_filled_cell(cell, value, target_kind.get(str(item.get("target") or ""), ""))
            filled += 1

        table_index = {str(t.get("target") or ""): t for t in list(schema.get("tables") or []) if isinstance(t, dict)}
        for item in list(plan.get("append_rows") or []):
            if not isinstance(item, dict):
                continue
            table = table_index.get(str(item.get("target") or ""))
            values = dict(item.get("values") or {}) if isinstance(item.get("values"), dict) else {}
            if not table or not values:
                continue
            try:
                ws = wb[str(table.get("sheet") or "")]
            except Exception:
                skipped += 1
                continue
            header_row = int(table.get("header_row") or 1)
            headers = [str(h or "").strip() for h in list(table.get("headers") or [])]
            append_row = max(ws.max_row + 1, header_row + 1)
            _copy_row_style(ws, header_row + 1 if ws.max_row >= header_row + 1 else header_row, append_row, len(headers))
            for idx, header in enumerate(headers, start=1):
                if header in values:
                    cell = ws.cell(append_row, idx)
                    if isinstance(cell, MergedCell):
                        skipped += 1
                        continue
                    cell.value = values[header]
            appended += 1
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell, MergedCell) or not isinstance(cell.value, str):
                        continue
                    text = str(cell.value or "")
                    labels = [str(x or "").strip() for x in re.findall(r"\{\{\s*([^{}]{1,50})\s*\}\}", text)]
                    if not labels:
                        continue
                    new_text = re.sub(r"\{\{\s*[^{}]{1,50}\s*\}\}", "", text).strip()
                    cell.value = new_text
                    cleared_placeholders += len(labels)
                    unresolved.extend([label for label in labels if label])
        out = io.BytesIO()
        wb.save(out)
    finally:
        try:
            wb.close()
        except Exception:
            pass
    return FillResult(
        file_bytes=out.getvalue(),
        stats={
            "filled_cells": filled,
            "appended_rows": appended,
            "skipped": skipped,
            "cleared_placeholders": cleared_placeholders,
            "planned_fills": len(list(plan.get("fills") or [])),
            "planned_append_rows": len(list(plan.get("append_rows") or [])),
        },
        warnings=warnings + [f"No value provided for '{label}'; placeholder cleared." for label in unresolved[:20]],
    )
