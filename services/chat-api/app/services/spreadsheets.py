from __future__ import annotations

from io import BytesIO
import re
from typing import Any, Dict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.enterprise_capabilities.spreadsheets.models import SpreadsheetColumnSpec, SpreadsheetSpec
from app.services.spreadsheet_validation import validate_rendered_workbook, validate_spreadsheet_spec
from app.utils.oss_uploader import AliyunOSSUploader


XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _safe_filename(value: str, ext: str = ".xlsx") -> str:
    cleaned = re.sub(r"[\\/:*?\"<>|\r\n\t]+", "_", str(value or "").strip())
    cleaned = re.sub(r"\s+", "_", cleaned).strip("_") or "spreadsheet"
    if not cleaned.lower().endswith(ext):
        cleaned = f"{cleaned}{ext}"
    return cleaned[:120]


class SpreadsheetService:
    def _render_validated(self, spec: SpreadsheetSpec) -> tuple[bytes, Dict[str, Any]]:
        validate_spreadsheet_spec(spec)
        wb = Workbook()
        default = wb.active
        wb.remove(default)

        for sheet_spec in spec.sheets:
            ws = wb.create_sheet(title=sheet_spec.name or "Sheet1")
            columns = list(sheet_spec.columns or [])
            rows = list(sheet_spec.rows or [])
            if not columns and rows:
                keys = list(rows[0].keys())
                columns = [
                    SpreadsheetColumnSpec(key=key, title=key, type="text")
                    for key in keys
                ]

            header_fill = PatternFill("solid", fgColor="E8EEF7")
            header_font = Font(bold=True, color="1F2937")
            wrap = Alignment(wrap_text=True, vertical="top")

            for col_idx, col in enumerate(columns, start=1):
                cell = ws.cell(row=1, column=col_idx, value=col.title)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = wrap

            for row_idx, row in enumerate(rows, start=2):
                for col_idx, col in enumerate(columns, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=row.get(col.key))
                    cell.alignment = wrap
                    if col.number_format:
                        cell.number_format = col.number_format
                    elif col.type == "currency":
                        cell.number_format = "#,##0.00"
                    elif col.type == "percent":
                        cell.number_format = "0.00%"
                    elif col.type == "date":
                        cell.number_format = "yyyy-mm-dd"

            if columns:
                ws.freeze_panes = "A2" if sheet_spec.formatting.get("freeze_header", True) else None
                if sheet_spec.formatting.get("autofilter", True):
                    ws.auto_filter.ref = ws.dimensions
                for col_idx, col in enumerate(columns, start=1):
                    letter = get_column_letter(col_idx)
                    if col.width:
                        ws.column_dimensions[letter].width = col.width
                        continue
                    values = [str(col.title or "")]
                    for row in rows[:100]:
                        values.append(str(row.get(col.key) or ""))
                    ws.column_dimensions[letter].width = max(10, min(50, max(len(v) for v in values) + 2))

        bio = BytesIO()
        wb.save(bio)
        content = bio.getvalue()
        validation = validate_rendered_workbook(content, spec)
        return content, validation

    def render_bytes(self, spec: SpreadsheetSpec) -> bytes:
        content, _validation = self._render_validated(spec)
        return content

    async def render(self, spec: SpreadsheetSpec, *, user_id: str, filename: str | None = None) -> Dict[str, Any]:
        file_name = filename or _safe_filename(spec.workbook_title)
        content, validation = self._render_validated(spec)
        uploader = AliyunOSSUploader()
        _, object_path = uploader.upload_bytes_with_path(
            content,
            user_id,
            file_name,
            content_type=XLSX_CONTENT_TYPE,
        )
        signed_url = uploader.sign_url(object_path)
        return {
            "type": "xlsx",
            "url": signed_url,
            "filename": file_name,
            "title": spec.workbook_title or file_name,
            "object_path": object_path,
            "content_type": XLSX_CONTENT_TYPE,
            "sheet_count": len(spec.sheets),
            "validation": validation,
        }


spreadsheet_service = SpreadsheetService()
