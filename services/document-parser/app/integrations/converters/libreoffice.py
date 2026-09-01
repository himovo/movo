from __future__ import annotations

import shutil
import subprocess
from copy import copy
from pathlib import Path
from typing import Any

from app.core.config import settings


def find_soffice() -> str:
    if settings.libreoffice_bin:
        return settings.libreoffice_bin
    for command in ("soffice", "libreoffice"):
        path = shutil.which(command)
        if path:
            return path
    mac_path = Path("/Applications/LibreOffice.app/Contents/MacOS/soffice")
    if mac_path.exists() and mac_path.is_file():
        return str(mac_path)
    raise RuntimeError("LibreOffice executable not found")


SPREADSHEET_EXTENSIONS = {".xlsx", ".xlsm"}
SPREADSHEET_PREVIEW_DEFAULT_ROW_HEIGHT = 18
SPREADSHEET_PREVIEW_MAX_ROW_HEIGHT = 24
SPREADSHEET_PREVIEW_MAX_COLUMN_WIDTH = 80
LEGACY_OFFICE_TARGET_FORMATS = {
    ".doc": ("docx", ".docx"),
    ".ppt": ("pptx", ".pptx"),
    ".xls": ("xlsx", ".xlsx"),
}


def convert_legacy_office_to_modern(source_path: Path, output_dir: Path) -> Path:
    source_suffix = source_path.suffix.lower()
    target = LEGACY_OFFICE_TARGET_FORMATS.get(source_suffix)
    if not target:
        return source_path

    target_format, target_suffix = target
    output_dir.mkdir(parents=True, exist_ok=True)
    profile_dir = output_dir / ".libreoffice-profile"
    profile_dir.mkdir(parents=True, exist_ok=True)
    command = [
        find_soffice(),
        "--headless",
        f"-env:UserInstallation={profile_dir.resolve().as_uri()}",
        "--convert-to",
        target_format,
        "--outdir",
        str(output_dir),
        str(source_path),
    ]
    result = subprocess.run(
        command,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        timeout=settings.conversion_timeout_seconds,
        check=False,
    )
    if result.returncode != 0:
        raise RuntimeError(result.stderr or result.stdout or "LibreOffice legacy format conversion failed")

    candidate = output_dir / f"{source_path.stem}{target_suffix}"
    if candidate.exists():
        return candidate

    candidates = sorted(output_dir.glob(f"*{target_suffix}"), key=lambda item: item.stat().st_mtime, reverse=True)
    if candidates:
        return candidates[0]
    raise RuntimeError(f"Converted {target_suffix} file was not generated")


def convert_office_to_pdf(source_path: Path, output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    conversion_source = _prepare_source_for_pdf_preview(source_path, output_dir)
    profile_dir = output_dir / ".libreoffice-profile"
    profile_dir.mkdir(parents=True, exist_ok=True)
    command = [
        find_soffice(),
        "--headless",
        f"-env:UserInstallation={profile_dir.resolve().as_uri()}",
        "--convert-to",
        "pdf",
        "--outdir",
        str(output_dir),
        str(conversion_source),
    ]
    result = subprocess.run(
        command,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        timeout=settings.conversion_timeout_seconds,
        check=False,
    )
    if result.returncode != 0:
        raise RuntimeError(result.stderr or result.stdout or "LibreOffice conversion failed")
    candidate = output_dir / f"{conversion_source.stem}.pdf"
    if not candidate.exists():
        raise RuntimeError("Converted PDF was not generated")
    return candidate


def _prepare_source_for_pdf_preview(source_path: Path, output_dir: Path) -> Path:
    if source_path.suffix.lower() not in SPREADSHEET_EXTENSIONS:
        return source_path
    return _prepare_spreadsheet_for_pdf_preview(source_path, output_dir)


def _prepare_spreadsheet_for_pdf_preview(source_path: Path, output_dir: Path) -> Path:
    try:
        from openpyxl import load_workbook  # type: ignore
        from openpyxl.utils import get_column_letter  # type: ignore
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to prepare spreadsheet PDF previews") from exc

    workbook = load_workbook(source_path, keep_vba=source_path.suffix.lower() == ".xlsm")
    for sheet in workbook.worksheets:
        if sheet.sheet_state != "visible":
            continue
        max_row, max_column = _used_range(sheet)
        if max_row <= 0 or max_column <= 0:
            continue

        _normalize_spreadsheet_preview_layout(sheet, max_row, max_column)
        sheet.print_area = f"A1:{get_column_letter(max_column)}{max_row}"
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A3
        sheet.page_margins.left = 0.25
        sheet.page_margins.right = 0.25
        sheet.page_margins.top = 0.35
        sheet.page_margins.bottom = 0.35
        if max_row >= 1:
            sheet.print_title_rows = "1:1"
        try:
            sheet.page_setup.scale = None
        except Exception:
            pass

    prepared_path = output_dir / f"{source_path.stem}.preview{source_path.suffix}"
    workbook.save(prepared_path)
    return prepared_path


def _normalize_spreadsheet_preview_layout(sheet: Any, max_row: int, max_column: int) -> None:
    # The generated PDF is used as an evidence locator, not as a print-ready
    # workbook. Normalize a temporary copy so rows stay in a grid-like preview
    # instead of expanding long wrapped cells across many PDF pages.
    sheet.row_breaks.brk = []
    sheet.col_breaks.brk = []

    for row_idx in range(1, max_row + 1):
        row_dimension = sheet.row_dimensions[row_idx]
        if row_dimension.hidden:
            continue
        if row_dimension.height is None:
            row_dimension.height = SPREADSHEET_PREVIEW_DEFAULT_ROW_HEIGHT
        else:
            row_dimension.height = min(
                float(row_dimension.height),
                SPREADSHEET_PREVIEW_MAX_ROW_HEIGHT,
            )

    for col_idx in range(1, max_column + 1):
        column_letter = sheet.cell(row=1, column=col_idx).column_letter
        column_dimension = sheet.column_dimensions[column_letter]
        if column_dimension.hidden:
            continue
        if column_dimension.width:
            column_dimension.width = min(
                float(column_dimension.width),
                SPREADSHEET_PREVIEW_MAX_COLUMN_WIDTH,
            )

    for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_column):
        for cell in row:
            alignment = copy(cell.alignment)
            alignment.wrap_text = False
            alignment.shrink_to_fit = False
            cell.alignment = alignment


def _used_range(sheet: Any) -> tuple[int, int]:
    max_row = 0
    max_column = 0
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            if isinstance(cell.value, str) and not cell.value.strip():
                continue
            max_row = max(max_row, int(cell.row))
            max_column = max(max_column, int(cell.column))
    return max_row, max_column
